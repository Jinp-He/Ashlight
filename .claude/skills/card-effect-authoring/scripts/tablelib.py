# -*- coding: utf-8 -*-
"""Luban xlsx 读写共享逻辑。把《excel-gotchas.md》里的坑一次性固化在这里：
- 数据行 A 列本就是空的（Id 在 B 列），判数据行绝不能用 A 列非空
- A 列 '##' 开头 = Luban 注释行（未导入游戏）
- 一个 workbook 里所有 A1=='##var' 的 sheet 合并成一张逻辑表
- 逻辑 sheet 名与物理 sheetN.xml 不保证同序，必须用 openpyxl
改这个文件时别把以上修复改回去。
"""
import io
import sys

import openpyxl

# Windows 控制台默认 GBK，统一 UTF-8 输出
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
if hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")


def read_sheet(xlsx_path):
    """合并所有 ##var sheet，返回字符串行列表（空单元格为 ''）。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    merged = []
    for ws in wb.worksheets:
        rows = [["" if v is None else str(v) for v in row]
                for row in ws.iter_rows(values_only=True)]
        if rows and rows[0] and rows[0][0].strip() == "##var":
            merged.extend(rows)
    wb.close()
    return merged


def header_columns(rows):
    """##var 行 -> {列名: 列号}。"""
    for row in rows:
        if row and row[0] == "##var":
            return {name: i for i, name in enumerate(row) if name}
    return {}


def data_rows(rows, include_off=False):
    """产出 (is_off, row)。is_off=True 表示 A 列 '##' 注释行（未导入游戏）。
    跳过整行空行和 ##var/##type/##group 元信息行。"""
    for row in rows:
        if not any((c or "").strip() for c in row):
            continue
        a = row[0].strip()
        if a.startswith("##var") or a.startswith("##type") or a.startswith("##group"):
            continue
        is_off = a.startswith("##")
        if is_off and not include_off:
            continue
        yield is_off, row


def collect_ids(xlsx_path, id_col_name="Id"):
    """启用行（非 ## 注释）的 Id 集合。"""
    rows = read_sheet(xlsx_path)
    ci = header_columns(rows).get(id_col_name)
    if ci is None:
        return set()
    return {row[ci] for _, row in data_rows(rows) if ci < len(row) and row[ci]}


def locate_cell(workbook, target_id, col_name):
    """在可写 workbook 中按 Id 定位单元格，返回 (worksheet, row, col)；找不到返回 None。
    跳过 ## 注释行；遍历所有 ##var sheet。"""
    for ws in workbook.worksheets:
        if ws["A1"].value != "##var":
            continue
        header = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
        id_col, col = header.get("Id"), header.get(col_name)
        if id_col is None or col is None:
            continue
        for r in range(2, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is not None and str(a).strip().startswith("##"):
                continue
            if ws.cell(row=r, column=id_col).value == target_id:
                return ws, r, col
    return None
