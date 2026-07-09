# -*- coding: utf-8 -*-
"""按 Id 定位改一个单元格，带旧值断言护栏（不匹配即中止不保存）。

用法（仓库根目录；改多个单元格就跑多次）：
    python .claude/skills/card-effect-authoring/scripts/edit_cell.py \
        "DataTables/Datas/Character/#CardInfo.xlsx" Zhouzhou010 CastZone Any Back

旧值必须精确等于当前单元格内容（不知道旧值先跑 dump_tables.py）。
Excel 开着目标文件时保存会 PermissionError，先关掉。
"""
import argparse
import sys

import openpyxl

from tablelib import locate_cell


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("id")
    ap.add_argument("column")
    ap.add_argument("old")
    ap.add_argument("new")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    hit = locate_cell(wb, args.id, args.column)
    if hit is None:
        print(f"中止：{args.xlsx} 里找不到 Id={args.id}（或无 {args.column} 列）", file=sys.stderr)
        return 1
    ws, r, c = hit
    cur = ws.cell(row=r, column=c).value
    if cur != args.old and str(cur) != args.old:
        print(f"中止：{args.id}.{args.column} 当前值 {cur!r} != 期望旧值 {args.old!r}", file=sys.stderr)
        return 1

    # 原单元格是数值类型且新值可解析时保持数值类型，避免 int 列被写成文本
    new = args.new
    if isinstance(cur, (int, float)) and not isinstance(cur, bool):
        try:
            new = type(cur)(args.new)
        except ValueError:
            pass
    ws.cell(row=r, column=c).value = new
    wb.save(args.xlsx)
    print(f"OK {args.id}.{args.column}: {cur!r} -> {new!r}  ({args.xlsx})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
