#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Static validation for character card authoring workbooks.

Run before Luban generation. Exit code 1 means the source table contains a
combination that is known to produce misleading text or dead configuration.
"""

from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CARD_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CardInfo.xlsx"
CHARACTER_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CharaterInfo.xlsx"
ROLE_PREFIXES = ("Rocket", "Irene", "Zhouzhou")
EXPECTED_TRAITS = {
    "Rocket": "FirstPushFree",
    "Irene": "DoubleExecution",
    "Zhouzhou": "MoveAllyGrantMorale",
}


def load_rows(path: Path, sheet: str, start_row: int = 5):
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb[sheet]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(header) if name}
    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start_row):
        if not any(value is not None for value in row):
            continue
        rows.append((row_number, row, index))
    return rows


def bool_value(value) -> bool:
    return str(value).strip().lower() in {"true", "1"}


def validate() -> list[str]:
    errors: list[str] = []
    cards = {}
    role_rarities = defaultdict(set)

    for sheet in ("Rocket", "Extra"):
        for row_number, row, index in load_rows(CARD_PATH, sheet):
            card_id = row[index["Id"]]
            if not card_id or str(card_id) == "string" or row[0] == "##":
                continue
            card_id = str(card_id)
            if card_id in cards:
                errors.append(f"duplicate Id {card_id} ({cards[card_id][0]} and {sheet}!R{row_number})")
                continue
            cards[card_id] = (f"{sheet}!R{row_number}", row, index)

            if not card_id.startswith(ROLE_PREFIXES):
                continue

            role = next(prefix for prefix in ROLE_PREFIXES if card_id.startswith(prefix))
            role_rarities[role].add(row[index["Rarity"]])
            card_type = row[index["CardType"]]
            executing_cost = int(row[index["ExecutingCost"]] or 0)
            description = str(row[index["Description"]] or "")
            effect_cells = [row[index["Effects"]]]
            for phase_field in ("ChargeStartEffects", "ChargeWhileEffects"):
                if phase_field in index and index[phase_field] < len(row):
                    effect_cells.append(row[index[phase_field]])
            effects = ";".join(str(value) for value in effect_cells if value)

            if card_type == "Swift" and executing_cost != 0:
                errors.append(f"{card_id}: Swift card has ExecutingCost={executing_cost}")
            if card_type == "Execution" and executing_cost <= 0:
                errors.append(f"{card_id}: Execution card must have ExecutingCost > 0")
            if card_type == "Charge" and executing_cost != 0:
                errors.append(f"{card_id}: Charge card must have ExecutingCost=0")

            effect_parts = [part.split(",") for part in effects.split(";") if part]
            effect_names = {part[0] for part in effect_parts if part}
            single_cast_effects = {
                "CastShiftEffect",
                "CastDamageBonusEffect",
                "CastResolveDrawEffect",
                "CastResolveBuffEffect",
                "CastImmediateEffect",
                "CastEchoEffect",
            }
            if row[index["TargetType"]] == "TimeSlot":
                if card_type != "Swift":
                    errors.append(f"{card_id}: TimeSlot target card must be Swift")
                if not (effect_names & single_cast_effects):
                    errors.append(f"{card_id}: TimeSlot target has no pending-cast effect")
            elif effect_names & single_cast_effects:
                errors.append(f"{card_id}: single pending-cast effect requires TargetType=TimeSlot")

            effect_code_counts = Counter(part[1] for part in effect_parts if len(part) >= 2)
            for code, raw_index in re.findall(r"\{([A-Z])(?::(\d+))?\}", description):
                normalized = "F" if code == "V" else code
                requested = int(raw_index or 0)
                if effect_code_counts[normalized] <= requested:
                    errors.append(
                        f"{card_id}: placeholder {{{code}{':' + raw_index if raw_index else ''}}} "
                        f"has no matching effect"
                    )

            multi_hit = re.search(r"(?:造成|连续造成)(\d+)次", description)
            if multi_hit:
                expected = int(multi_hit.group(1))
                attack_count = sum(part[0] == "AttackEffect" for part in effect_parts)
                if attack_count != expected:
                    errors.append(f"{card_id}: text says {expected} hits but has {attack_count} AttackEffect entries")

            for part in effect_parts:
                if part[0] == "PushCollisionEffect" and len(part) >= 4:
                    collision = part[3]
                    if role != "Rocket" and collision != "None":
                        errors.append(f"{card_id}: non-warrior card must not own collision result {collision}")

            if card_id.endswith("000") and bool_value(row[index["IsInUpgrade"]]):
                errors.append(f"{card_id}: injected temporary move card cannot enter reward pool")

    for role, rarities in role_rarities.items():
        if not {"普通", "稀有", "史诗"}.issubset(rarities):
            errors.append(f"{role}: rarity tiers incomplete: {sorted(rarities)}")

    if "Extra001" in cards:
        _, row, index = cards["Extra001"]
        if bool_value(row[index["IsInUpgrade"]]):
            errors.append("Extra001: generated knife cannot enter reward pool")

    character_rows = load_rows(CHARACTER_PATH, "Sheet1", start_row=4)
    referenced_cards = set()
    seen_traits = {}
    for row_number, row, index in character_rows:
        character = row[index["Character"]]
        if character not in EXPECTED_TRAITS:
            continue
        seen_traits[character] = row[index["Trait"]]
        referenced_cards.update(str(row[index["BaseDeck"]] or "").split(","))

    for character, expected in EXPECTED_TRAITS.items():
        if seen_traits.get(character) != expected:
            errors.append(f"{character}: Trait must be {expected}, got {seen_traits.get(character)!r}")
    for card_id in referenced_cards:
        if card_id and card_id not in cards:
            errors.append(f"base deck references missing card {card_id}")

    return errors


def main() -> int:
    errors = validate()
    if errors:
        print("Character card validation failed:")
        for error in errors:
            print(f"  - {error}")
        return 1
    print("Character card validation passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
