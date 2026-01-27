# -*- coding: utf-8 -*-
"""
卡牌数值平衡 - Excel 修改脚本
直接修改 #CardInfo.xlsx 中的 Effects 列数值
"""

from openpyxl import load_workbook
import os
import shutil
import re

# Excel 文件路径
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Datas", "Character", "#CardInfo.xlsx")

# 平衡数据：卡牌ID -> 需要修改的数值
# Effects 格式: "EffectType,字段1,字段2,..."
# 例如: "AttackEffect,A,6,false" 中的 6 是伤害值
BALANCE_UPDATES = {
    # ========== Rocket (战士) - 2格技能 ==========
    "Rocket001": {"damage": "10"},           # 打击: 6→10
    "Rocket002": {"damage": "8"},            # 迅猛打击: 6→8
    "Rocket003": {"damage": "8"},            # 旋风斩: 6→8
    "Rocket004": {"damage": "8"},            # 巨龙撞击: 6→8
    "Rocket005": {"shield": "20"},           # 守护: 30→20
    "Rocket007": {"heal": "8"},              # 吞噬痛苦: 10→8
    "Rocket009": {"damage": "8", "bonus": "8"},  # 交错打击: 6+6→8+8

    # ========== Irene (法师) - 2格技能 ==========
    "Irene001": {"damage": "10"},            # 火球术: 30→10
    "Irene002": {"damage": "8"},             # 冰霜震击: 10→8
    "Irene003": {"damage": "4"},             # 奥术飞弹: 5→4 (每次)
    "Irene004": {"defense": "10"},           # 寒冰护甲: 15→10
    "Irene005": {"damage": "25"},            # 炎爆术: 80→25
    "Irene006": {"damage": "12"},            # 暴风雪: 20→12
    "Irene009": {"damage": "25"},            # 陨石灭世: 100→25
    "Irene011": {"damage": "10"},            # 镜像实体: 20→10

    # ========== Zhouzhou (游侠) - 3格技能 ==========
    "Zhouzhou001": {"damage": "12"},         # 快速射击: 15→12
    "Zhouzhou002": {"damage": "12"},         # 强力击: 20→12
    "Zhouzhou003": {"damage": "8"},          # 二连矢: 12→8 (每次)
    "Zhouzhou004": {"damage": "12"},         # 多重射击: 15→12
    "Zhouzhou005": {"damage": "25"},         # 狙击: 60→25
    "Zhouzhou006": {"damage": "8"},          # 后撤步: 5→8
    "Zhouzhou008": {"damage": "12", "bonus": "12"},  # 致命射击: 25+25→12+12
    "Zhouzhou012": {"damage": "18"},         # 穿云箭: 40→18
}


def update_effects_value(effects_str, card_id, updates):
    """
    更新 Effects 字符串中的数值

    Effects 格式示例:
    - "AttackEffect,A,6,false" -> 伤害在第3个位置
    - "AttackExtraEffect,A,6,Channeling|Recoit,2.0" -> 伤害在第3个位置
    - "DefenseEffect,D,3,true" -> 护甲在第3个位置
    - "HealEffect,H,10" -> 治疗在第3个位置
    - "InterceptEffect,D,30" -> 护盾在第3个位置
    - "AttackConditionalEffect,A,6,IsAttacking" -> 条件伤害在第3个位置
    """
    if not effects_str or not updates:
        return effects_str

    # 分割多个效果 (用分号分隔)
    effect_parts = effects_str.split(';')
    modified_parts = []

    for part in effect_parts:
        # 分割单个效果的字段 (用逗号分隔)
        fields = part.split(',')
        if len(fields) < 2:
            modified_parts.append(part)
            continue

        effect_type = fields[0]
        modified = False

        # 根据效果类型更新对应数值
        if effect_type == "AttackEffect" and "damage" in updates:
            # AttackEffect,A,damage,is_aoe
            if len(fields) >= 3:
                fields[2] = updates["damage"]
                modified = True

        elif effect_type == "AttackExtraEffect" and "damage" in updates:
            # AttackExtraEffect,A,damage,conditions,multiplier
            if len(fields) >= 3:
                fields[2] = updates["damage"]
                modified = True

        elif effect_type == "AttackConditionalEffect" and "bonus" in updates:
            # AttackConditionalEffect,A,bonus_damage,condition_type
            if len(fields) >= 3:
                fields[2] = updates["bonus"]
                modified = True

        elif effect_type == "DefenseEffect" and "defense" in updates:
            # DefenseEffect,D,value,per_hit
            if len(fields) >= 3:
                fields[2] = updates["defense"]
                modified = True

        elif effect_type == "HealEffect" and "heal" in updates:
            # HealEffect,H,value
            if len(fields) >= 3:
                fields[2] = updates["heal"]
                modified = True

        elif effect_type == "InterceptEffect" and "shield" in updates:
            # InterceptEffect,D,shield_value
            if len(fields) >= 3:
                fields[2] = updates["shield"]
                modified = True

        if modified:
            modified_parts.append(','.join(fields))
        else:
            modified_parts.append(part)

    return ';'.join(modified_parts)


def main():
    print("=" * 60)
    print("卡牌数值平衡 - Excel 修改工具")
    print("=" * 60)

    # 检查文件
    excel_path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(excel_path):
        print(f"错误: 未找到文件 {excel_path}")
        return

    print(f"目标文件: {excel_path}")

    # 备份
    backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
    shutil.copy2(excel_path, backup_path)
    print(f"已备份到: {backup_path}")

    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到列索引
    id_col = 2       # B列 = Id
    effects_col = 5  # E列 = Effects

    # 统计
    updated_count = 0

    print("\n" + "-" * 60)
    print("开始更新...")
    print("-" * 60)

    # 遍历数据行
    for row in range(1, ws.max_row + 1):
        card_id = ws.cell(row=row, column=id_col).value

        if card_id and card_id in BALANCE_UPDATES:
            old_effects = ws.cell(row=row, column=effects_col).value
            updates = BALANCE_UPDATES[card_id]

            new_effects = update_effects_value(old_effects, card_id, updates)

            if old_effects != new_effects:
                ws.cell(row=row, column=effects_col, value=new_effects)
                print(f"\n{card_id}:")
                print(f"  旧: {old_effects}")
                print(f"  新: {new_effects}")
                updated_count += 1

    # 保存
    if updated_count > 0:
        wb.save(excel_path)
        print("\n" + "=" * 60)
        print(f"完成! 共更新 {updated_count} 张卡牌")
        print("=" * 60)
        print("\n请运行 gen.bat 重新生成配置文件")
    else:
        print("\n没有需要更新的内容")


if __name__ == "__main__":
    main()
