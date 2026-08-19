#!/usr/bin/env python3
"""Normalize Rocket cards in #CardInfo.xlsx without touching other card rows."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CARD_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CardInfo.xlsx"
CHARACTER_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CharaterInfo.xlsx"
ENUM_PATH = ROOT / "DataTables" / "Datas" / "__enums__.xlsx"

EFFECT_LIST_TYPE = "(list#sep=;),(Effect#sep=,)"


FIELDS = (
    "Name",
    "Description",
    "Effects",
    "ChargeStartEffects",
    "ChargeWhileEffects",
    "CardType",
    "BelongTo",
    "TargetType",
    "Rarity",
    "AssetPath",
    "IsLocked",
    "Energy",
    "ExecutingCost",
    "IsEthereal",
    "IsExhaust",
    "TargetZone",
    "CastZone",
    "IsInUpgrade",
)


def card(
    name: str,
    description: str,
    effects: str | None,
    card_type: str,
    target_type: str,
    rarity: str,
    energy: int,
    executing_cost: int = 0,
    *,
    target_zone: str = "Any",
    cast_zone: str = "Any",
    is_locked: str = "FALSE",
    is_in_upgrade: str = "TRUE",
    is_ethereal: str = "FALSE",
    is_exhaust: str = "FALSE",
    charge_start_effects: str | None = None,
    charge_while_effects: str | None = None,
):
    return {
        "Name": name,
        "Description": description,
        "Effects": effects,
        "ChargeStartEffects": charge_start_effects,
        "ChargeWhileEffects": charge_while_effects,
        "CardType": card_type,
        "BelongTo": "Rocket",
        "TargetType": target_type,
        "Rarity": rarity,
        "AssetPath": None,
        "IsLocked": is_locked,
        "Energy": energy,
        "ExecutingCost": executing_cost,
        "IsEthereal": is_ethereal,
        "IsExhaust": is_exhaust,
        "TargetZone": target_zone,
        "CastZone": cast_zone,
        "IsInUpgrade": is_in_upgrade,
    }


CARDS = {
    "Rocket001": card(
        "挥砍", "造成{A}点伤害。", "AttackEffect,A,7,false",
        "Swift", "SingleEnemy", "基础", 1, is_in_upgrade="FALSE",
    ),
    "Rocket002": card(
        "重型护板", "获得{D}点护甲。", "DefenseEffect,D,9,false",
        "Swift", "Self", "稀有", 2,
    ),
    "Rocket003": card(
        "炮击", "对所有敌人造成{A}点伤害。", "AttackEffect,A,6,true",
        "Swift", "AllEnemy", "稀有", 2,
    ),
    "Rocket004": card(
        "超载火炮", "[过载]{V}。对[前排]所有敌人造成{A}点伤害。",
        "OverloadEffect,F,1;AttackEffect,A,10,true",
        "Swift", "AllEnemy", "稀有", 2, target_zone="Front",
    ),
    "Rocket005": card(
        "举盾", "获得{D}点护甲并移动至[前排]。",
        "DefenseEffect,D,6,false;MoveSelfEffect,N,FrontRow",
        "Swift", "Self", "基础", 1, is_in_upgrade="FALSE",
    ),
    "Rocket006": card(
        "挺身而出", "[蓄力]：立即获得{D}点护甲；蓄力期间获得[嘲讽]。",
        None,
        "Charge", "Self", "史诗", 0,
        charge_start_effects="DefenseEffect,D,10,false",
        charge_while_effects="TauntEffect,N,All",
    ),
    "Rocket007": card(
        "奇思妙想", "随机获得{N}张[小发明]；若自身位于[后排]，获得数量翻倍。",
        "AddRandomToHandEffect,N,Extra002|Extra003|Extra004,1,SelfInBackRow,2",
        "Swift", "Self", "普通", 1,
    ),
    "Rocket008": card(
        "护甲包", "使一名友方角色获得{D}点护甲。", "DefenseEffect,D,8,false",
        "Swift", "SingleAlly", "稀有", 2,
    ),
    "Rocket009": card(
        "交错打击", "造成{A}点伤害；若有友方角色在本轮行动，额外造成{A:1}点伤害。",
        "AttackEffect,A,9,false;AttackConditionalEffect,A,9,AllyActingThisRound",
        "Swift", "SingleEnemy", "普通", 2,
    ),
    "Rocket010": card(
        "重炮", "[过载]{V}。[蓄力]：造成{A}点伤害。",
        "OverloadEffect,F,1;AttackEffect,A,20,false",
        "Charge", "SingleEnemy", "基础", 0,
        cast_zone="Back", is_in_upgrade="FALSE",
    ),
    "Rocket011": card(
        "装甲齐射", "对[前排]所有敌人造成{A}点伤害；每击中一名敌人，获得{D}点护甲。",
        "AttackEffect,A,5,true;DefenseEffect,D,2,true",
        "Swift", "AllEnemy", "基础", 2,
        target_zone="Front", is_in_upgrade="FALSE",
    ),
    "Rocket012": card(
        "阵线加固", "使[前排]所有友方角色获得{D}点护甲。",
        "DefenseEffect,D,6,false",
        "Swift", "AllAlly", "基础", 2,
        target_zone="Front", is_in_upgrade="FALSE",
    ),
    "Rocket013": card(
        "超频供能", "[过载]{V}。获得{V:1}点能量，抽{V:2}张牌。",
        "OverloadEffect,F,2;EnergyEffect,F,2;DrawEffect,F,1",
        "Swift", "Self", "基础", 0, is_in_upgrade="FALSE",
    ),
    "Rocket014": card(
        "极限超频", "[过载]{V}。获得{V:1}点能量，抽{V:2}张牌。",
        "OverloadEffect,F,3;EnergyEffect,F,3;DrawEffect,F,2",
        "Swift", "Self", "史诗", 0,
    ),
    "Rocket015": card(
        "冷却循环", "清除自身所有[过载]，抽{V}张牌。",
        "ClearOverloadEffect,N;DrawEffect,F,1",
        "Swift", "Self", "稀有", 1,
    ),
    "Rocket016": card(
        "应急装甲", "[蓄力]：获得{D}点护甲；护甲耗尽时移动至[后排]。",
        "DefenseEffect,D,12,false;MoveOnArmorBreakEffect,N,BackRow",
        "Charge", "Self", "稀有", 0,
    ),
    "Rocket017": card(
        "震荡冲击", "[过载]{V}。造成{A}点伤害并[推迟]{T}回合；若发生[碰撞]，碰撞双方[晕眩]。",
        "OverloadEffect,F,1;AttackEffect,A,12,false;PushCollisionEffect,T,1,Stun",
        "Swift", "SingleEnemy", "稀有", 2,
    ),
    "Rocket018": card(
        "战术打击", "[蓄力]：造成{A}×蓄力层数点伤害。",
        "ChargedAttackEffect,A,5,false",
        "Charge", "SingleEnemy", "史诗", 0,
    ),
    "Rocket019": card(
        "区域装甲", "使与自己同排的所有友方角色获得{D}点护甲。",
        "DefenseEffect,D,7,false",
        "Swift", "AllAlly", "稀有", 2,
        target_zone="Conditional",
    ),
    "Rocket020": card(
        "推进阵线", "移动至[前排]，自己与前排所有队友各获得{D}点护甲。",
        "MoveSelfEffect,N,FrontRow;DefenseEffect,D,5,false",
        "Swift", "AllAlly", "稀有", 2, target_zone="Front",
    ),
    "Rocket000": card(
        "移动", "[移动]。", "MovePositionEffect,N,Toggle",
        "Swift", "Self", "基础", 1,
        is_in_upgrade="FALSE", is_ethereal="TRUE", is_exhaust="TRUE",
    ),
}


EXTRA_CARDS = {
    "Extra002": {
        "Name": "折叠护板",
        "Description": "获得{D}点护甲。",
        "Effects": "DefenseEffect,D,4,false",
        "TargetType": "Self",
    },
    "Extra003": {
        "Name": "袖珍火箭",
        "Description": "造成{A}点伤害。",
        "Effects": "AttackEffect,A,5,false",
        "TargetType": "SingleEnemy",
    },
    "Extra004": {
        "Name": "快速装填",
        "Description": "抽{V}张牌。",
        "Effects": "DrawEffect,F,1",
        "TargetType": "Self",
    },
}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        src = ws.cell(source_row, column)
        dst = ws.cell(target_row, column)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def find_id_rows(ws, id_column: int) -> dict[str, int]:
    result = {}
    for row in range(5, ws.max_row + 1):
        value = ws.cell(row, id_column).value
        if value:
            result[str(value)] = row
    return result


def ensure_card_effect_columns(ws) -> dict[str, int]:
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    for field in ("ChargeStartEffects", "ChargeWhileEffects"):
        if field in headers:
            continue
        column = ws.max_column + 1
        source_column = headers["Effects"]
        for row in range(1, ws.max_row + 1):
            src = ws.cell(row, source_column)
            dst = ws.cell(row, column)
            if src.has_style:
                dst._style = copy(src._style)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)
        ws.cell(1, column).value = field
        ws.cell(2, column).value = EFFECT_LIST_TYPE
        ws.cell(3, column).value = None
        ws.cell(4, column).value = None
        headers[field] = column
    return headers


def ensure_charge_enum() -> None:
    wb = load_workbook(ENUM_PATH)
    ws = wb.active
    card_type_row = next(
        row for row in range(1, ws.max_row + 1)
        if ws.cell(row, 2).value == "CardTypeEnum"
    )
    row = card_type_row + 1
    while row <= ws.max_row and ws.cell(row, 2).value in (None, ""):
        name = ws.cell(row, 8).value
        if name == "Charge":
            return
        if name in (None, ""):
            break
        row += 1

    ws.insert_rows(row)
    source_row = row - 1
    for column in range(1, ws.max_column + 1):
        src = ws.cell(source_row, column)
        dst = ws.cell(row, column)
        if src.has_style:
            dst._style = copy(src._style)
        if src.alignment:
            dst.alignment = copy(src.alignment)
    ws.cell(row, 8).value = "Charge"
    ws.cell(row, 9).value = "蓄力行动"
    wb.save(ENUM_PATH)


def main() -> None:
    ensure_charge_enum()
    wb = load_workbook(CARD_PATH)
    ws = wb["Rocket"]
    headers = ensure_card_effect_columns(ws)
    id_rows = find_id_rows(ws, headers["Id"])

    missing = sorted(set(CARDS) - set(id_rows))
    if missing:
        raise RuntimeError(f"Rocket rows are missing: {missing}")

    for card_id, values in CARDS.items():
        row = id_rows[card_id]
        # Every normalized Rocket entry is an authored card, not a commented
        # template row. Some legacy rows (notably Rocket009) still carried ##.
        ws.cell(row, 1).value = None
        for field in FIELDS:
            ws.cell(row, headers[field]).value = values[field]
        effect_cell = ws.cell(row, headers["Effects"])
        effect_cell.comment = None

    # Luban treats a whitespace-only cell on an otherwise empty row as data.
    # Remove only whitespace remnants from rows that have neither an Id nor a
    # comment marker, leaving all authored and commented-out card rows intact.
    for row in range(5, ws.max_row + 1):
        marker = ws.cell(row, 1).value
        card_id = ws.cell(row, headers["Id"]).value
        if marker is not None or card_id is not None:
            continue
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row, column)
            if isinstance(cell.value, str) and not cell.value.strip():
                cell.value = None

    extra = wb["Extra"]
    extra_headers = ensure_card_effect_columns(extra)
    extra_rows = find_id_rows(extra, extra_headers["Id"])
    for extra_id, authored in EXTRA_CARDS.items():
        extra_row = extra_rows.get(extra_id)
        if extra_row is None:
            extra_row = extra.max_row + 1
            copy_row_style(extra, 5, extra_row)
            extra_rows[extra_id] = extra_row

        extra_values = {
            "Id": extra_id,
            "Name": authored["Name"],
            "Description": authored["Description"],
            "Effects": authored["Effects"],
            "ChargeStartEffects": None,
            "ChargeWhileEffects": None,
            "CardType": "Swift",
            "BelongTo": "Rocket",
            "TargetType": authored["TargetType"],
            "Rarity": "临时",
            "AssetPath": None,
            "IsLocked": "FALSE",
            "Energy": 0,
            "ExecutingCost": 0,
            "IsEthereal": "TRUE",
            "IsExhaust": "TRUE",
            "TargetZone": "Any",
            "CastZone": "Any",
            "IsInUpgrade": "FALSE",
        }
        for field, value in extra_values.items():
            extra.cell(extra_row, extra_headers[field]).value = value

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(CARD_PATH)

    character_book = load_workbook(CHARACTER_PATH)
    character_sheet = character_book["Sheet1"]
    character_headers = {cell.value: cell.column for cell in character_sheet[1] if cell.value}
    for row in range(4, character_sheet.max_row + 1):
        if character_sheet.cell(row, character_headers["Character"]).value == "Rocket":
            character_sheet.cell(row, character_headers["BaseDeck"]).value = (
                "Rocket001,Rocket001,Rocket001,Rocket001,Rocket001,"
                "Rocket005,Rocket005,Rocket005,Rocket005,Rocket005"
            )
            break
    else:
        raise RuntimeError("Rocket character row is missing")
    if character_book.calculation is not None:
        character_book.calculation.fullCalcOnLoad = True
        character_book.calculation.forceFullCalc = True
    character_book.save(CHARACTER_PATH)

    print(
        f"Normalized {len(CARDS)} Rocket cards and "
        f"{len(EXTRA_CARDS)} inventions in {CARD_PATH} and updated Rocket base deck"
    )


if __name__ == "__main__":
    main()
