#!/usr/bin/env python3
"""Normalize Irene cards in #CardInfo.xlsx without touching other card rows."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CARD_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CardInfo.xlsx"
CHARACTER_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CharaterInfo.xlsx"


FIELDS = (
    "Name",
    "Description",
    "Effects",
    "ChargeStartEffects",
    "ChargeWhileEffects",
    "CardType",
    "BelongTo",
    "TargetType",
    "Rarity",
    "AssetPath",
    "IsLocked",
    "Energy",
    "ExecutingCost",
    "IsEthereal",
    "IsExhaust",
    "TargetZone",
    "CastZone",
    "IsInUpgrade",
)


def card(
    name: str,
    description: str,
    effects: str,
    card_type: str,
    target_type: str,
    rarity: str,
    energy: int,
    executing_cost: int = 0,
    *,
    target_zone: str = "Any",
    cast_zone: str = "Any",
    is_in_upgrade: str = "TRUE",
    is_ethereal: str = "FALSE",
    is_exhaust: str = "FALSE",
):
    return {
        "Name": name,
        "Description": description,
        "Effects": effects,
        "ChargeStartEffects": None,
        "ChargeWhileEffects": None,
        "CardType": card_type,
        "BelongTo": "Irene",
        "TargetType": target_type,
        "Rarity": rarity,
        "AssetPath": None,
        "IsLocked": "FALSE",
        "Energy": energy,
        "ExecutingCost": executing_cost,
        "IsEthereal": is_ethereal,
        "IsExhaust": is_exhaust,
        "TargetZone": target_zone,
        "CastZone": cast_zone,
        "IsInUpgrade": is_in_upgrade,
    }


CARDS = {
    "Irene001": card(
        "火球术", "造成{A}点伤害。", "AttackEffect,A,18,false",
        "Execution", "SingleEnemy", "基础", 2, 3, is_in_upgrade="FALSE",
    ),
    "Irene002": card(
        "奥术护罩", "使[前排]所有友方角色获得{D}点护甲。", "DefenseEffect,D,9,false",
        "Execution", "AllAlly", "基础", 3, 2,
        target_zone="Front", is_in_upgrade="FALSE",
    ),
    "Irene003": card(
        "秘法飞弹", "连续造成3次{A}点伤害。",
        "AttackEffect,A,3,false;AttackEffect,A,3,false;AttackEffect,A,3,false",
        "Swift", "SingleEnemy", "基础", 1, is_in_upgrade="FALSE",
    ),
    "Irene004": card(
        "引导", "抽{V}张牌。", "DrawEffect,F,1",
        "Swift", "Self", "基础", 0, is_in_upgrade="FALSE",
    ),
    "Irene005": card(
        "炎爆术", "造成{A}点伤害。", "AttackEffect,A,28,false",
        "Execution", "SingleEnemy", "普通", 3, 4,
    ),
    "Irene006": card(
        "寒霜新星", "对[前排]所有敌人造成{A}点伤害，并使其行动[推迟]{T}格。",
        "AttackEffect,A,10,true;TimeShiftAllEffect,T,2",
        "Execution", "AllEnemy", "基础", 4, 4,
        target_zone="Front", is_in_upgrade="FALSE",
    ),
    "Irene007": card(
        "静滞", "造成{A}点伤害，并使目标[定身]。",
        "AttackEffect,A,10,false;BuffEffect,F,Root,1",
        "Execution", "SingleEnemy", "基础", 2, 2, is_in_upgrade="FALSE",
    ),
    "Irene008": card(
        "闪现", "移动一名友方角色，使其获得{D}点护甲。",
        "MovePositionEffect,N,Toggle;DefenseEffect,D,3,false",
        "Swift", "SingleAlly", "基础", 1, is_in_upgrade="FALSE",
    ),
    "Irene009": card(
        "连锁闪电", "对所有敌人造成{A}点伤害；若这个公共回合结算过天气，伤害翻倍。",
        "WeatherConditionalAttackEffect,A,10,2",
        "Execution", "AllEnemy", "稀有", 2, 3,
    ),
    "Irene010": card(
        "余烬增压", "选择一张会造成伤害的己方在轨执行牌，使其[顺延]{T}格；该牌造成的伤害增加{A}点。",
        "CastShiftEffect,T,2;CastDamageBonusEffect,A,8",
        "Swift", "TimeSlot", "基础", 0,
        is_in_upgrade="FALSE", is_exhaust="TRUE",
    ),
    "Irene011": card(
        "时间复利", "选择一张己方在轨执行牌，使其[顺延]{T}格；该牌结算时，抽{V}张牌，并使其施法者获得{V:1}点[充能]。",
        "CastShiftEffect,T,2;CastResolveDrawEffect,F,2;CastResolveBuffEffect,F,Energized,1",
        "Swift", "TimeSlot", "稀有", 1, is_exhaust="TRUE",
    ),
    "Irene012": card(
        "抢拍", "选择一张己方在轨执行牌，使其[提前]{T}格；[过载]{V}。",
        "CastShiftEffect,T,-1;OverloadEffect,F,1",
        "Swift", "TimeSlot", "基础", 0,
        is_in_upgrade="FALSE", is_exhaust="TRUE",
    ),
    "Irene013": card(
        "提前兑现", "选择一张己方在轨执行牌，使其在当前公共回合立刻结算；[过载]{V}。",
        "CastImmediateEffect,N;OverloadEffect,F,2",
        "Swift", "TimeSlot", "稀有", 1, is_exhaust="TRUE",
    ),
    "Irene014": card(
        "回声咏唱", "选择一张含有伤害、护甲或治疗效果的己方在轨执行牌，使其[顺延]{T}格；该牌首次结算后，在{V}个公共回合后以一半数值再次结算这些效果。",
        "CastShiftEffect,T,3;CastEchoEffect,F,2,0.5",
        "Swift", "TimeSlot", "史诗", 1, is_exhaust="TRUE",
    ),
    "Irene015": card(
        "时缝", "使一名敌人的行动[推迟]{T}格。",
        "PushCollisionEffect,T,2,None",
        "Swift", "SingleEnemy", "普通", 1,
    ),
    "Irene016": card(
        "拽引", "使一名敌人的行动[提前]{T}格。",
        "PushCollisionEffect,T,-2,None",
        "Swift", "SingleEnemy", "普通", 1,
    ),
    "Irene017": card(
        "时间馈赠", "使一名友方角色的行动[提前]{T}格，并使其获得{V}点[充能]。",
        "PushCollisionEffect,T,-1,None;BuffEffect,F,Energized,1",
        "Swift", "SingleAlly", "稀有", 2,
    ),
    "Irene018": card(
        "观测天象", "抽{V}张牌；若有己方在轨执行牌将在下一次天气所在的公共回合结算，获得{V:1}点能量。",
        "DrawEffect,F,1;WeatherSyncEnergyEffect,F,1",
        "Swift", "Self", "普通", 0, is_exhaust="TRUE",
    ),
    "Irene019": card(
        "引雷", "使一名敌人的下一次行动向下一次天气所在的公共回合移动至多{T}格，但不会越过该公共回合。",
        "AlignToWeatherEffect,T,2",
        "Swift", "SingleEnemy", "稀有", 1,
    ),
    "Irene020": card(
        "终焉倒数", "造成{A}点伤害；此牌每被己方卡牌[顺延]，伤害便提高，最多翻倍。",
        "DelayScaledAttackEffect,A,24,6,24,false",
        "Execution", "SingleEnemy", "史诗", 4, 6,
    ),
    "Irene021": card(
        "风暴眼", "直到下一次天气结算，安排在该公共回合行动的所有友方角色在天气结算前获得{D}点护甲。",
        "WeatherGuardEffect,D,8",
        "Swift", "Self", "稀有", 2, is_exhaust="TRUE",
    ),
    "Irene022": card(
        "时间透支", "[过载]{V}。获得{V:1}点能量，抽{V:2}张牌。",
        "OverloadEffect,F,2;EnergyEffect,F,2;DrawEffect,F,1",
        "Swift", "Self", "稀有", 0,
    ),
    "Irene023": card(
        "改天换日", "使下一次天气[顺延]{T}格；使所有己方在轨执行牌[提前]{T:1}格。",
        "WeatherShiftEffect,T,3;CastShiftAllEffect,T,-1",
        "Swift", "Self", "史诗", 2, is_exhaust="TRUE",
    ),
    "Irene000": card(
        "移动", "[移动]。", "MovePositionEffect,N,Toggle",
        "Swift", "Self", "基础", 1,
        is_in_upgrade="FALSE", is_ethereal="TRUE", is_exhaust="TRUE",
    ),
}


def main() -> None:
    workbook = load_workbook(CARD_PATH)
    worksheet = workbook["Rocket"]
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value}
    missing_headers = sorted((set(FIELDS) | {"Id"}) - set(headers))
    if missing_headers:
        raise RuntimeError(f"Card table headers are missing: {missing_headers}")

    id_rows: dict[str, int] = {}
    duplicate_rows: list[int] = []
    for row in range(5, worksheet.max_row + 1):
        value = worksheet.cell(row, headers["Id"]).value
        if not value:
            continue
        card_id = str(value)
        if card_id in CARDS and card_id in id_rows:
            duplicate_rows.append(row)
            continue
        id_rows[card_id] = row
    missing_cards = sorted(set(CARDS) - set(id_rows))
    if missing_cards:
        raise RuntimeError(f"Irene rows are missing: {missing_cards}")

    for card_id, values in CARDS.items():
        row = id_rows[card_id]
        worksheet.cell(row, 1).value = None
        for field in FIELDS:
            worksheet.cell(row, headers[field]).value = values[field]
        worksheet.cell(row, headers["Effects"]).comment = None

    # Remove only duplicate Irene data rows, keeping their formatting intact.
    for row in duplicate_rows:
        for column in range(1, worksheet.max_column + 1):
            worksheet.cell(row, column).value = None

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(CARD_PATH)

    character_workbook = load_workbook(CHARACTER_PATH)
    character_sheet = character_workbook["Sheet1"]
    character_headers = {cell.value: cell.column for cell in character_sheet[1] if cell.value}
    for row in range(4, character_sheet.max_row + 1):
        if character_sheet.cell(row, character_headers["Character"]).value == "Irene":
            character_sheet.cell(row, character_headers["BaseDeck"]).value = (
                "Irene001,Irene001,Irene002,Irene006,Irene007,"
                "Irene003,Irene004,Irene008,Irene010,Irene012"
            )
            break
    else:
        raise RuntimeError("Irene character row is missing")
    if character_workbook.calculation is not None:
        character_workbook.calculation.fullCalcOnLoad = True
        character_workbook.calculation.forceFullCalc = True
    character_workbook.save(CHARACTER_PATH)

    print(f"Normalized {len(CARDS)} Irene cards in {CARD_PATH} and updated Irene base deck")


if __name__ == "__main__":
    main()
