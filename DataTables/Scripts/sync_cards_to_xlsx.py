# -*- coding: utf-8 -*-
"""
将卡牌平衡改动 + 新增卡同步回 #CardInfo.xlsx
（运行 gen.bat 时 Luban 会从该 xlsx 重新生成 JSON，避免改动被覆盖）
"""
import os
import shutil
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Datas", "Character", "#CardInfo.xlsx")
# 备份目录必须在 Datas/ 之外，否则 Luban 会把备份当成新表
BACKUP_DIR = os.path.join(os.path.dirname(__file__), "..", "_backups")

# 13 处数值修正：card_id -> 新的 Effects 字符串
BALANCE_UPDATES = {
    "Rocket003":   "AttackEffect,A,11,true;DefenseEffect,D,3,true",
    "Rocket004":   "AttackEffect,A,14,false;PushCollisionEffect,T,1,Stun",
    "Rocket005":   "InterceptEffect,D,14",
    "Irene001":    "AttackEffect,A,14,false",
    "Irene002":    "AttackEffect,A,11,false;PushCollisionEffect,T,1,None",
    "Irene004":    "DefenseEffect,D,8,false;BuffEffect,F,Chill,1",
    "Irene006":    "AttackEffect,A,18,true;TimeShiftAllEffect,T,1",
    "Irene009":    "AttackExtraEffect,A,16,Channeling|Recoil,2.0",
    "Irene010":    "AttackEffect,A,28,true;TimeShiftAllEffect,T,2;BuffEffect,F,Chill,2",
    "Irene011":    "AttackEffect,A,4,false;AttackConditionalEffect,A,6,InExecution",
    "Zhouzhou002": "AttackEffect,A,15,false;PushCollisionEffect,T,2,Stun",
    "Zhouzhou004": "AttackEffect,A,14,true",
    "Zhouzhou007": "AttackEffect,A,14,false;BuffEffect,F,Root,1",
}

# 12 张新卡（顺序与 BelongTo 一致）
# 字段顺序: Id, Name, Description, Effects, CardType, BelongTo, TargetType, Rarity, AssetPath, IsLocked, Energy, ExecutingCost
NEW_CARDS = [
    ("Rocket010", "破韧斩",
     "造成 {A} 点伤害，并对目标施加 [破韧] {V}（累计承受 {V} 点伤害后行动被打断）。",
     "AttackEffect,A,8,false;BuffEffect,F,Stagger,12",
     "Execution", "Rocket", "SingleEnemy", "普通", None, "FALSE", "1", 2),
    ("Rocket011", "战吼",
     "获得 [力量] {V}（攻击伤害 +{V}，持续 2 回合）。",
     "BuffEffect,F,Strength,3",
     "Swift", "Rocket", "Self", "普通", None, "FALSE", "1", 0),
    ("Rocket012", "盾墙之姿",
     "获得 {D} 点护甲，并获得 [敏捷] {V}（防御牌额外 +{V} 护甲）。",
     "InterceptEffect,D,12;BuffEffect,F,Dexterity,3",
     "Execution", "Rocket", "Self", "普通", None, "FALSE", "1", 2),
    ("Rocket013", "粉碎一击",
     "造成 {A} 点伤害，并使目标 [晕眩] {V} 回合（打断其正在执行的行动）。",
     "AttackEffect,A,6,false;BuffEffect,F,Stun,1",
     "Execution", "Rocket", "SingleEnemy", "普通", None, "FALSE", "1", 3),

    ("Irene013", "灼热印记",
     "造成 {A} 点伤害，并施加 [燃烧] {V}（每回合开始受到 {V} 点伤害，持续 3 回合）。",
     "AttackEffect,A,6,false;BuffEffect,F,Burn,4",
     "Execution", "Irene", "SingleEnemy", "普通", None, "FALSE", "1", 1),
    ("Irene014", "法术屏障",
     "获得 [圣物] {V}（接下来 {V} 次 debuff 被自动抵消）。",
     "BuffEffect,F,Artifact,2",
     "Swift", "Irene", "Self", "普通", None, "FALSE", "1", 0),
    ("Irene015", "元素涌动",
     "获得 [充能] {V}（下回合开始 +{V} 能量）。",
     "BuffEffect,F,Energized,2",
     "Swift", "Irene", "Self", "普通", None, "FALSE", "1", 0),
    ("Irene016", "凋零诅咒",
     "对目标施加 [虚弱] {V}%（造成伤害降低）和 [脆弱] {V:1}%（受到的伤害提高）。",
     "BuffEffect,F,Weak,40;BuffEffect,F,Frail,50",
     "Execution", "Irene", "SingleEnemy", "普通", None, "FALSE", "1", 2),

    ("Zhouzhou013", "蛇毒矢",
     "造成 {A} 点伤害，并施加 [中毒] {V}（每回合开始受到 {V} 点伤害并衰减 1）。",
     "AttackEffect,A,6,false;BuffEffect,F,Poison,4",
     "Execution", "Zhouzhou", "SingleEnemy", "普通", None, "FALSE", "1", 2),
    ("Zhouzhou014", "闪避姿态",
     "获得 [敏捷] {V}（防御牌额外 +{V} 护甲，持续 2 回合）。",
     "BuffEffect,F,Dexterity,3",
     "Swift", "Zhouzhou", "Self", "普通", None, "FALSE", "1", 0),
    ("Zhouzhou015", "断筋箭",
     "造成 {A} 点伤害，施加 [脆弱] {V}%（获得护甲降低）和 [禁锢] {V:1}（无法移动）。",
     "AttackEffect,A,10,false;BuffEffect,F,Frail,50;BuffEffect,F,Root,1",
     "Execution", "Zhouzhou", "SingleEnemy", "普通", None, "FALSE", "1", 2),
    ("Zhouzhou016", "战术准备",
     "获得 [灵感] {V}（下回合多抽 {V} 张牌）和 [减费] {V:1}（下一张牌费用 -{V:1}）。",
     "BuffEffect,F,DrawCard,1;BuffEffect,F,CardCost,1",
     "Swift", "Zhouzhou", "Self", "普通", None, "FALSE", "1", 0),
]


def main():
    path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(path):
        print(f"ERR: 未找到 {path}")
        return

    os.makedirs(BACKUP_DIR, exist_ok=True)
    backup = os.path.join(os.path.abspath(BACKUP_DIR), os.path.basename(path).replace(".xlsx", "_backup_before_new_cards.xlsx"))
    shutil.copy2(path, backup)
    print(f"已备份到 {backup}")

    wb = load_workbook(path)
    ws = wb.active

    # 1. 更新 13 张卡的 Effects
    updated = []
    for row in range(1, ws.max_row + 1):
        cid = ws.cell(row=row, column=2).value
        if cid in BALANCE_UPDATES:
            old = ws.cell(row=row, column=5).value
            new = BALANCE_UPDATES[cid]
            if old != new:
                ws.cell(row=row, column=5, value=new)
                updated.append((cid, old, new))

    print(f"\n[1/2] 更新 {len(updated)} 张卡的 Effects:")
    for cid, old, new in updated:
        print(f"  {cid}:")
        print(f"    旧: {old}")
        print(f"    新: {new}")

    # 2. 追加 12 张新卡
    start_row = ws.max_row + 1
    print(f"\n[2/2] 追加 {len(NEW_CARDS)} 张新卡 (起始行 {start_row}):")
    for i, card in enumerate(NEW_CARDS):
        row = start_row + i
        # C1 留空（##var 行专用），C2-C13 写数据
        for col_idx, value in enumerate(card, start=2):
            ws.cell(row=row, column=col_idx, value=value)
        print(f"  R{row}: {card[0]} ({card[1]})")

    wb.save(path)
    print(f"\n保存完成: {path}")
    print("\n下一步: 运行 DataTables/gen.bat 重新生成 JSON")


if __name__ == "__main__":
    main()
