"""Configure the first-pass monastery and wilderness enemy tables.

The script is intentionally idempotent: matching IDs are updated in place and
missing IDs are appended.  It also keeps one pre-edit backup of each workbook.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "Datas"
ENEMY_DIR = DATA_DIR / "Enemy"
CARD_PATH = DATA_DIR / "Character" / "#CardInfo.xlsx"
BACKUP_DIR = ROOT / "_backups"


CARD_ROW = {
    "Id": "Extra006",
    "Name": "解签",
    "Description": "[诅咒]。此牌在手牌中时，其他卡牌费用+1。",
    "Effects": None,
    "CardType": "Swift",
    "BelongTo": "Irene",
    "TargetType": "Self",
    "Rarity": "临时",
    "Energy": 1,
    "ExecutingCost": 0,
    "IsEthereal": False,
    "IsExhaust": True,
    "TargetZone": "Any",
    "CastZone": "Any",
    "IsInUpgrade": False,
}


SKILL_ROWS = [
    {
        "Id": "ESM01",
        "Name": "杖击",
        "Description": "对前排一名角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,4,false",
        "ExecutingCost": 1,
        "TargetType": "SingleEnemy",
        "TargetZone": "Front",
    },
    {
        "Id": "ESM02",
        "Name": "塞签",
        "Description": "将1张[解签]加入目标角色的手牌。",
        "Effects": "AddToHandEffect,N,Extra006,1",
        "ExecutingCost": 2,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESM03",
        "Name": "摸索祷词",
        "Description": "对后排一名角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,6,false",
        "ExecutingCost": 2,
        "TargetType": "SingleEnemy",
        "TargetZone": "Back",
    },
    {
        "Id": "ESM04",
        "Name": "盲签训诫",
        "Description": "对后排一名角色造成{A}点伤害，将1张[解签]加入其手牌。",
        "Effects": "AttackEffect,A,3,false;AddToHandEffect,N,Extra006,1",
        "ExecutingCost": 3,
        "TargetType": "SingleEnemy",
        "TargetZone": "Back",
    },
    {
        "Id": "ESM05",
        "Name": "戒尺横扫",
        "Description": "对前排所有角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,7,true",
        "ExecutingCost": 3,
        "TargetType": "AllEnemy",
        "TargetZone": "Front",
    },
    {
        "Id": "ESM06",
        "Name": "钟下裁定",
        "Description": "对目标角色造成{A}点伤害，将1张[解签]加入其手牌；引导期间受到{B}点伤害即被打断。",
        "Effects": "AttackEffect,A,8,false;AddToHandEffect,N,Extra006,1;BuffEffect,B,Stagger,14",
        "ExecutingCost": 5,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESM07",
        "Name": "封口裁决",
        "Description": "对目标角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,10,false",
        "ExecutingCost": 2,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESM08",
        "Name": "戒律护身",
        "Description": "获得{D}点护甲和{B}层力量。",
        "Effects": "DefenseEffect,D,8,false;BuffEffect,B,Strength,1",
        "ExecutingCost": 2,
        "TargetType": "Self",
        "TargetZone": "Any",
    },
    {
        "Id": "ESM09",
        "Name": "大斋戒",
        "Description": "对目标角色造成{A}点伤害，将1张[解签]加入其手牌；引导期间受到{B}点伤害即被打断。",
        "Effects": "AttackEffect,A,12,false;AddToHandEffect,N,Extra006,1;BuffEffect,B,Stagger,18",
        "ExecutingCost": 5,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESW01",
        "Name": "弹指",
        "Description": "对前排一名角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,4,false",
        "ExecutingCost": 1,
        "TargetType": "SingleEnemy",
        "TargetZone": "Front",
    },
    {
        "Id": "ESW02",
        "Name": "掷骨",
        "Description": "对后排一名角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,4,false",
        "ExecutingCost": 1,
        "TargetType": "SingleEnemy",
        "TargetZone": "Back",
    },
    {
        "Id": "ESW03",
        "Name": "嗅血",
        "Description": "对目标角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,6,false",
        "ExecutingCost": 2,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESW04",
        "Name": "扑食",
        "Description": "对目标角色造成{A}点伤害。",
        "Effects": "AttackEffect,A,10,false",
        "ExecutingCost": 3,
        "TargetType": "SingleEnemy",
        "TargetZone": "Any",
    },
    {
        "Id": "ESW05",
        "Name": "轮臂",
        "Description": "对前排一名角色造成3次{A}点伤害。",
        "Effects": "AttackEffect,A,3,false;AttackEffect,A,3,false;AttackEffect,A,3,false",
        "ExecutingCost": 3,
        "TargetType": "SingleEnemy",
        "TargetZone": "Front",
    },
    {
        "Id": "ESW06",
        "Name": "献臂冲锋",
        "Description": "对后排所有角色造成{A}点伤害；引导期间受到{B}点伤害即被打断。",
        "Effects": "AttackEffect,A,12,true;BuffEffect,B,Stagger,16",
        "ExecutingCost": 5,
        "TargetType": "AllEnemy",
        "TargetZone": "Back",
    },
]


ENEMY_ROWS = [
    {
        "Id": "EnemyM01",
        "Name": "灰袍信徒",
        "AlternativePath": "AshrobeDevotee",
        "Hp": 36,
        "Speed": 2,
        "IntentionSet": "A0,ESM01,1;A1,ESM02,2",
        "IsElite": False,
        "StartRow": "Front",
    },
    {
        "Id": "EnemyM02",
        "Name": "蒙眼侍僧",
        "AlternativePath": "VeiledAcolyte",
        "Hp": 54,
        "Speed": 3,
        "IntentionSet": "A0,ESM03,2;A1,ESM04,3",
        "IsElite": False,
        "StartRow": "Back",
    },
    {
        "Id": "EnemyM03",
        "Name": "戒律执事",
        "AlternativePath": "TheDisciplinarian",
        "Hp": 72,
        "Speed": 4,
        "IntentionSet": "A0,ESM05,3;A1,ESM06,5",
        "IsElite": False,
        "StartRow": "Front",
    },
    {
        "Id": "EnemyM04",
        "Name": "缄默使徒",
        "AlternativePath": "Mafe",
        "Hp": 150,
        "Speed": 3,
        "IntentionSet": "A0,ESM07,2;A1,ESM08,2;A2,ESM09,5",
        "IsElite": True,
        "StartRow": "Front",
    },
    {
        "Id": "EnemyW01",
        "Name": "指裔",
        "AlternativePath": "Tusk",
        "Hp": 36,
        "Speed": 2,
        "IntentionSet": "A0,ESW01,1;A1,ESW02,1",
        "IsElite": False,
        "StartRow": "Front",
    },
    {
        "Id": "EnemyW02",
        "Name": "血肉猎手",
        "AlternativePath": "Wangg",
        "Hp": 54,
        "Speed": 3,
        "IntentionSet": "A0,ESW03,2;A1,ESW04,3",
        "IsElite": False,
        "StartRow": "Back",
    },
    {
        "Id": "EnemyW03",
        "Name": "多臂殉道者",
        "AlternativePath": "Mafe",
        "Hp": 72,
        "Speed": 4,
        "IntentionSet": "A0,ESW05,3;A1,ESW06,5",
        "IsElite": False,
        "StartRow": "Front",
    },
]


ENCOUNTER_ROWS = [
    {"Id": "M101", "EnemySet": "EnemyM01", "StrategySet": ["A0;1.0"]},
    {"Id": "M102", "EnemySet": "EnemyM01,EnemyM02", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "M103", "EnemySet": "EnemyM01,EnemyM03", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "M104", "EnemySet": "EnemyM02,EnemyM03", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "M105", "EnemySet": "EnemyM04,EnemyM01", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "W101", "EnemySet": "EnemyW01,EnemyW01", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "W102", "EnemySet": "EnemyW01,EnemyW02", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "W103", "EnemySet": "EnemyW01,EnemyW03", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "W104", "EnemySet": "EnemyW02,EnemyW03", "StrategySet": ["A0,A0;1.0"]},
    # Boundary intrusion templates: one same-tier slot has already been replaced.
    {"Id": "MW101", "EnemySet": "EnemyM03,EnemyW01", "StrategySet": ["A0,A0;1.0"]},
    {"Id": "WM101", "EnemySet": "EnemyW03,EnemyM01", "StrategySet": ["A0,A0;1.0"]},
]


LEGACY_ENCOUNTER_IDS = {
    "EM000": "M101",
    "EM001": "M102",
    "EM002": "M103",
    "EM003": "M104",
    "EME01": "M105",
    "EW000": "W101",
    "EW001": "W102",
    "EW002": "W103",
    "EW003": "W104",
    "EMW01": "MW101",
    "EWM01": "WM101",
}


def backup_once(source: Path, backup_name: str) -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / backup_name
    if not backup.exists():
        copy2(source, backup)


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def upsert_rows(path: Path, rows: list[dict], sheet_name: str = "Sheet1") -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = {
        str(ws.cell(1, column).value): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(1, column).value
    }
    id_column = headers["Id"]
    existing = {
        str(ws.cell(row, id_column).value): row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, id_column).value
    }
    style_row = max(existing.values())

    for values in rows:
        row = existing.get(values["Id"])
        if row is None:
            row = ws.max_row + 1
            copy_row_style(ws, style_row, row)
            existing[values["Id"]] = row
        for header, value in values.items():
            if header == "StrategySet":
                for offset, strategy in enumerate(value):
                    ws.cell(row, headers["StrategySet"] + offset, strategy)
                continue
            ws.cell(row, headers[header], value)
    wb.save(path)


def migrate_encounter_ids(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    existing = {
        str(ws.cell(row, 2).value): row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 2).value
    }
    for old_id, new_id in LEGACY_ENCOUNTER_IDS.items():
        if old_id in existing and new_id not in existing:
            ws.cell(existing[old_id], 2, new_id)
    wb.save(path)


def validate_source_tables() -> None:
    card_wb = load_workbook(CARD_PATH, read_only=True, data_only=False)
    card_ids = {
        row[1]
        for sheet in card_wb.worksheets
        for row in sheet.iter_rows(values_only=True)
        if len(row) > 1 and row[1] and not str(row[1]).startswith("##")
    }
    assert "Extra006" in card_ids

    def ids(path: Path) -> set[str]:
        ws = load_workbook(path, read_only=True, data_only=False)["Sheet1"]
        return {
            str(row[1])
            for row in ws.iter_rows(values_only=True)
            if len(row) > 1 and row[1] and not str(row[1]).startswith("##")
        }

    enemy_ids = ids(ENEMY_DIR / "#EnemyInfo.xlsx")
    skill_ids = ids(ENEMY_DIR / "#EnemySkillInfo.xlsx")
    encounter_ids = ids(ENEMY_DIR / "#Encounter.xlsx")
    assert {row["Id"] for row in ENEMY_ROWS} <= enemy_ids
    assert {row["Id"] for row in SKILL_ROWS} <= skill_ids
    assert {row["Id"] for row in ENCOUNTER_ROWS} <= encounter_ids


def main() -> None:
    backup_once(
        ENEMY_DIR / "#EnemyInfo.xlsx",
        "#EnemyInfo_backup_before_monastery_wilderness.xlsx",
    )
    backup_once(
        ENEMY_DIR / "#EnemySkillInfo.xlsx",
        "#EnemySkillInfo_backup_before_monastery_wilderness.xlsx",
    )
    backup_once(
        ENEMY_DIR / "#Encounter.xlsx",
        "#Encounter_backup_before_monastery_wilderness.xlsx",
    )

    # Extra006 normally already exists; this keeps the card definition reproducible.
    upsert_rows(CARD_PATH, [CARD_ROW], sheet_name="Extra")
    upsert_rows(ENEMY_DIR / "#EnemySkillInfo.xlsx", SKILL_ROWS)
    upsert_rows(ENEMY_DIR / "#EnemyInfo.xlsx", ENEMY_ROWS)
    migrate_encounter_ids(ENEMY_DIR / "#Encounter.xlsx")
    upsert_rows(ENEMY_DIR / "#Encounter.xlsx", ENCOUNTER_ROWS)
    validate_source_tables()

    print(f"card: {CARD_ROW['Id']}")
    print(f"skills: {len(SKILL_ROWS)}")
    print(f"enemies: {len(ENEMY_ROWS)}")
    print(f"encounters: {len(ENCOUNTER_ROWS)}")
    print("source table validation: passed")


if __name__ == "__main__":
    main()
