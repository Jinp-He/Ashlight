"""Creates the editable source workbook for map tile content weights."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "DataTables" / "Datas" / "Map" / "#MapTileContentWeight.xlsx"

HEADERS = [
    "Shape",
    "EmptyRoadWeight",
    "BattleWeight",
    "EventWeight",
    "EliteWeight",
    "BattleEncounterId",
    "EliteEncounterId",
]

ROWS = [
    ["End", 65, 15, 18, 2, "E001", "E101"],
    ["Straight", 45, 30, 20, 5, "E001", "E101"],
    ["Turn", 45, 25, 25, 5, "E001", "E101"],
    ["TShaped", 30, 35, 20, 15, "E001", "E101"],
    ["Cross", 20, 40, 20, 20, "E001", "E101"],
]


def main() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MapTileContentWeight"
    sheet.sheet_view.showGridLines = False

    sheet.append(["##var", *HEADERS])
    sheet.append(["##type", "string", "int", "int", "int", "int", "string", "string"])
    sheet.append(["##group", "c", "c", "c", "c", "c", "c", "c"])
    sheet.append(["##comment", "道路形状", "空路权重", "普通敌人权重", "事件权重", "精英敌人权重", "普通敌人遭遇 ID", "精英敌人遭遇 ID"])
    for row in ROWS:
        sheet.append([None, *row])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for cell in sheet[2]:
        cell.font = Font(color="666666", italic=True)
    for cell in sheet[4]:
        cell.font = Font(color="666666", italic=True)
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=2, max_col=8):
        for cell in row:
            cell.fill = input_fill

    for column, width in {"A": 14, "B": 16, "C": 18, "D": 16, "E": 16, "F": 18, "G": 22, "H": 20}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B5"
    sheet.auto_filter.ref = "A1:H9"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)


if __name__ == "__main__":
    main()
