"""
Add new enemies / skills / encounters to enemy xlsx tables.

Schema reminders (see __beans__.xlsx):
  Effect          base bean (var EffectType, sep=',')
  AttackEffect    Effect + (Damage:int, ToAll:bool), sep=','
  DefenseEffect   Effect + (Defense:int, PerHit:bool), sep=','
  BuffEffect      Effect + (buff_id:string, value:float), sep=','
  EnemyIntention  fields (EnemyIntentionType, EnemySkillIndex, TimeSlot), sep=','
  EnemyIntentions wrapper (EnemyIntentionList: list#sep=| of EnemyIntention), sep=';'

EffectEnum values: A=Attack, D=Defense, T=TimeSlot, H=Heal, N=Null, B=Buff, F=Fill
TargetTypeEnum:    SingleAlly=0, AllAlly=1, Self=2, SingleEnemy=3, AllEnemy=4, TimeSlot=5
EnemyIntentionEnum: A0/A1/A2/S0/S1

EnemyInfo IntentionSet column format (sep=';' between EnemyIntentions,
sep='|' between EnemyIntention inside one group, sep=',' between fields):
  A0,ES001,3   = one group with one intention
  A0,ES001,3;A1,ES002,5 = two separate groups, one intention each
  A0,ES001,3|A1,ES002,5 = single group with two intentions (rarely used now)

EnemySkillInfo Effects column format (sep=';' between effects, sep=',' inside):
  AttackEffect,A,6,false                   = single Attack 6 damage, single target
  AttackEffect,A,5,true;BuffEffect,B,Burn,3 = aoe atk + buff
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

ENEMY_DIR = Path(r"F:/Ashlight/DataTables/Datas/Enemy")

# ---------- EnemySkillInfo ----------
# columns: Id, Name, Description, Effects, ExecutingCost, TargetType
new_skills = [
    ("ES003", "铁锤重击",     "蓄力一击，造成{A}点伤害",                "AttackEffect,A,18,false",                              8,  "SingleEnemy"),
    ("ES004", "烈焰扫射",     "对所有目标造成{A}点伤害",                  "AttackEffect,A,5,true",                                5,  "AllEnemy"),
    ("ES005", "铁壁加固",     "获得{D}点护甲，并减伤{V}%",                "DefenseEffect,D,10,false;BuffEffect,B,ReduceDmg,30",   2,  "Self"),
    ("ES006", "急袭斩",       "快速一击，造成{A}点伤害",                  "AttackEffect,A,4,false",                               1,  "SingleEnemy"),
    ("ES007", "寒冰之握",     "造成{A}点伤害，对目标施加冰冻",             "AttackEffect,A,3,false;BuffEffect,B,Chill,1",          3,  "SingleEnemy"),
    ("ES008", "终极燃烧",     "对所有目标造成{A}点伤害，并施加燃烧",       "AttackEffect,A,8,true;BuffEffect,B,Burn,3",            10, "AllEnemy"),
    ("ES009", "战吼",         "获得{V}层力量",                            "BuffEffect,B,Strength,2",                              3,  "Self"),
    ("ES010", "易伤诅咒",     "对目标施加{V}%易伤",                       "BuffEffect,B,Vulnerable,50",                           3,  "SingleEnemy"),
]

# ---------- EnemyInfo ----------
# columns: Id, Name, AlternativePath, Hp, Speed, IntentionSet
# IntentionSet groups separated by ';'
new_enemies = [
    ("Enemy003", "重装步兵", "DogKnight", 140, 4, "A0,ES005,3;A1,ES003,8;A2,ES001,3"),
    ("Enemy004", "冰焰法师", "DogKnight", 65,  5, "A0,ES007,4;A1,ES004,5;S0,ES008,10"),
    ("Enemy005", "急袭兵",   "DogKnight", 55,  7, "A0,ES006,2;S0,ES009,3;S1,ES010,3"),
]

# ---------- Encounter ----------
# columns: Id, EnemySet, then per-strategy cells (list of Strategy in successive cols)
# Strategy format inside one cell: "<EnemyIntentionsScript>;<Weight>"
#   where EnemyIntentionsScript is "A0,A0,..." matching EnemySet order
new_encounters = [
    ("E002", "Enemy001,Enemy005",            [("A0,A0", 1.0)]),
    ("E003", "Enemy003,Enemy004",            [("A0,A0", 0.3), ("A1,A1", 0.4), ("A2,S0", 0.3)]),
    ("E004", "Enemy003,Enemy001,Enemy005",   [("A0,A0,A0", 0.5), ("A1,A0,S0", 0.5)]),
]


def find_first_blank_row(ws, header_col=2):
    """Return the row index after the last data row (cell in column header_col)."""
    last = 0
    for r in range(1, ws.max_row + 2):
        v = ws.cell(row=r, column=header_col).value
        if v is not None and not str(v).startswith("##"):
            last = r
    return last + 1 if last else ws.max_row + 1


def append_skill_rows():
    path = ENEMY_DIR / "#EnemySkillInfo.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    start = find_first_blank_row(ws)
    for i, (sid, name, desc, eff, cost, tgt) in enumerate(new_skills):
        r = start + i
        ws.cell(row=r, column=1, value=None)  # ##var column blank
        ws.cell(row=r, column=2, value=sid)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=eff)
        ws.cell(row=r, column=6, value=cost)
        ws.cell(row=r, column=7, value=tgt)
    wb.save(path)
    print(f"[skills] appended {len(new_skills)} rows starting at row {start}")


def append_enemy_rows():
    path = ENEMY_DIR / "#EnemyInfo.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    start = find_first_blank_row(ws)
    for i, (eid, name, alt, hp, spd, intent) in enumerate(new_enemies):
        r = start + i
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=eid)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=alt)
        ws.cell(row=r, column=5, value=hp)
        ws.cell(row=r, column=6, value=spd)
        ws.cell(row=r, column=7, value=intent)
    wb.save(path)
    print(f"[enemies] appended {len(new_enemies)} rows starting at row {start}")


def append_encounter_rows():
    path = ENEMY_DIR / "#Encounter.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    start = find_first_blank_row(ws)
    for i, (eid, enemy_set, strategies) in enumerate(new_encounters):
        r = start + i
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=eid)
        ws.cell(row=r, column=3, value=enemy_set)
        # strategies start at column 4, one Strategy per cell
        for j, (script, weight) in enumerate(strategies):
            ws.cell(row=r, column=4 + j, value=f"{script};{weight}")
    wb.save(path)
    print(f"[encounters] appended {len(new_encounters)} rows starting at row {start}")


if __name__ == "__main__":
    append_skill_rows()
    append_enemy_rows()
    append_encounter_rows()
    print("done")
