#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Apply the 2026-07 character-card audit fixes to the Luban source workbooks.

The script is intentionally idempotent. It preserves workbook layout/styles and
creates one recoverable pre-fix backup outside Datas/ so Luban will not ingest it.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CARD_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CardInfo.xlsx"
CHARACTER_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CharaterInfo.xlsx"
BACKUP_DIR = ROOT / "DataTables" / "_backups"


CARD_UPDATES = {
    "Rocket003": {
        "Description": "对所有敌人造成{A}点伤害，每命中一名敌人，获得{D}点护甲。",
        "Effects": "AttackEffect,A,6,true;DefenseEffect,D,2,true",
        "Rarity": "稀有",
    },
    "Rocket004": {
        "Description": "造成{A}点伤害并[推迟]{T}格。若与另一名敌人发生[碰撞]，双方[晕眩]1回合。",
        "Effects": "AttackEffect,A,8,false;PushCollisionEffect,T,1,Stun",
        "Rarity": "稀有",
    },
    "Rocket006": {
        "Description": "[执行]3。结算时[嘲讽]全体，并获得50%[减伤]，持续2回合。",
        "CardType": "Execution",
        "ExecutingCost": 3,
        "Rarity": "史诗",
    },
    "Rocket014": {
        "Description": "造成{A}点伤害并[推迟]{T}格。若与另一名敌人发生[碰撞]，双方[晕眩]1回合。",
        "Energy": 2,
        "Rarity": "稀有",
    },
    "Rocket016": {
        "Description": "[移动]目标，并使其[推迟]{T}格。若与另一名敌人发生[碰撞]，双方[晕眩]1回合。",
        "Rarity": "稀有",
    },
    "Rocket017": {
        "Effects": "AttackEffect,A,24,false",
        "Rarity": "稀有",
    },
    "Rocket018": {"Rarity": "史诗"},

    "Irene003": {
        "Description": "连续造成3次{A}点伤害。",
        "Effects": "AttackEffect,A,3,false;AttackEffect,A,3,false;AttackEffect,A,3,false",
    },
    "Irene009": {
        "Description": "对所有敌人造成{A}点伤害。若目标处于[执行]中，伤害翻倍。",
        "Effects": "AttackExtraEffect,A,12,Channeling|Recoil,2.0",
        "Rarity": "稀有",
    },
    "Irene010": {"Rarity": "史诗"},
    "Irene012": {"Rarity": "稀有"},
    "Irene013": {
        "Effects": "PushCollisionEffect,T,2,None",
    },
    "Irene014": {
        "Effects": "PushCollisionEffect,T,-2,None",
        "Rarity": "稀有",
    },
    "Irene015": {
        "Effects": "PushCollisionEffect,T,-1,None;BuffEffect,F,Energized,1",
        "Energy": 2,
        "Rarity": "稀有",
    },
    "Irene017": {"Rarity": "稀有"},
    "Irene018": {"Energy": 3, "Rarity": "史诗"},
    "Irene019": {
        "Description": "获得{D}点护甲与[充能]{V}。",
        "Effects": "DefenseEffect,D,4,false;BuffEffect,F,Energized,1",
        "Rarity": "稀有",
    },
    "Irene020": {"Rarity": "史诗"},
    "Irene021": {"Rarity": "史诗"},
    "Irene022": {"Rarity": "稀有"},
    "Irene023": {"Rarity": "史诗"},

    "Zhouzhou001": {
        "Effects": "AttackEffect,A,4,false",
        "Energy": 0,
    },
    "Zhouzhou002": {"Energy": 0},
    "Zhouzhou003": {"Energy": 0},
    "Zhouzhou005": {
        "Effects": "AttackEffect,A,8,false;DefenseConditionalEffect,D,4,SelfInFrontRow",
        "Rarity": "稀有",
    },
    "Zhouzhou007": {"Rarity": "稀有"},
    "Zhouzhou009": {
        "Effects": "OverloadEffect,F,2;AttackEffect,A,4,true;AttackEffect,A,4,true;AttackEffect,A,4,true",
        "Energy": 0,
        "Rarity": "史诗",
    },
    "Zhouzhou010": {"Rarity": "稀有"},
    "Zhouzhou011": {"Rarity": "稀有"},
    "Zhouzhou012": {"Rarity": "稀有"},
    "Zhouzhou013": {
        "Effects": "AttackCurrentRoundEffect,A,4",
        "Energy": 0,
    },
    "Zhouzhou014": {"Rarity": "史诗"},
    "Zhouzhou015": {"Rarity": "稀有"},
    "Zhouzhou016": {"Rarity": "稀有"},
    "Zhouzhou018": {"Rarity": "稀有"},
    "Zhouzhou019": {"Rarity": "史诗"},
    "Zhouzhou020": {"Rarity": "稀有"},
}


STARTER_OR_TEMPORARY_IDS = {
    "Rocket000", "Rocket001", "Rocket005", "Rocket010", "Rocket011", "Rocket012", "Rocket013",
    "Irene000", "Irene001", "Irene002", "Irene003", "Irene004", "Irene006", "Irene007", "Irene008",
    "Zhouzhou000", "Zhouzhou003", "Zhouzhou004", "Zhouzhou006", "Zhouzhou010", "Zhouzhou013", "Zhouzhou015",
}


DEFAULT_RARITIES = {
    "Rocket002": "稀有", "Rocket007": "普通", "Rocket008": "稀有", "Rocket015": "稀有",
    "Irene005": "稀有", "Irene011": "普通", "Irene016": "普通",
    "Zhouzhou017": "普通",
}


TRAITS = {
    "Rocket": "FirstPushFree",
    "Irene": "DoubleExecution",
    "Zhouzhou": "MoveAllyGrantMorale",
}


def backup_once(path: Path) -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"{path.stem}_backup_before_full_card_audit_fix.xlsx"
    if not backup.exists():
        shutil.copy2(path, backup)


def headers_for(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def apply_card_updates() -> None:
    backup_once(CARD_PATH)
    wb = load_workbook(CARD_PATH)
    ws = wb["Rocket"]
    headers = headers_for(ws)
    seen = set()

    for row in range(5, ws.max_row + 1):
        card_id = ws.cell(row=row, column=headers["Id"]).value
        if not card_id or str(card_id).startswith("##"):
            continue
        card_id = str(card_id)
        if not card_id.startswith(("Rocket", "Irene", "Zhouzhou")):
            continue
        seen.add(card_id)

        # Swift cards resolve immediately; stale executing costs must never affect authoring or UI.
        if ws.cell(row=row, column=headers["CardType"]).value == "Swift":
            ws.cell(row=row, column=headers["ExecutingCost"], value=0)

        if card_id in CARD_UPDATES:
            for field, value in CARD_UPDATES[card_id].items():
                ws.cell(row=row, column=headers[field], value=value)

        if card_id in DEFAULT_RARITIES:
            ws.cell(row=row, column=headers["Rarity"], value=DEFAULT_RARITIES[card_id])

        ws.cell(
            row=row,
            column=headers["IsInUpgrade"],
            value="FALSE" if card_id in STARTER_OR_TEMPORARY_IDS else "TRUE",
        )

    missing = sorted(set(CARD_UPDATES) - seen)
    if missing:
        raise RuntimeError(f"Card ids not found: {missing}")

    extra = wb["Extra"]
    extra_headers = headers_for(extra)
    for row in range(5, extra.max_row + 1):
        if extra.cell(row=row, column=extra_headers["Id"]).value == "Extra001":
            extra.cell(row=row, column=extra_headers["IsInUpgrade"], value="FALSE")

    wb.save(CARD_PATH)


def apply_character_traits() -> None:
    backup_once(CHARACTER_PATH)
    wb = load_workbook(CHARACTER_PATH)
    ws = wb["Sheet1"]
    headers = headers_for(ws)
    seen = set()
    for row in range(4, ws.max_row + 1):
        character = ws.cell(row=row, column=headers["Character"]).value
        if character in TRAITS:
            ws.cell(row=row, column=headers["Trait"], value=TRAITS[character])
            seen.add(character)
    if seen != set(TRAITS):
        raise RuntimeError(f"Characters not found: {sorted(set(TRAITS) - seen)}")
    wb.save(CHARACTER_PATH)


def validate_written_workbooks() -> None:
    wb = load_workbook(CARD_PATH, data_only=False, read_only=True)
    ws = wb["Rocket"]
    headers = {cell.value: cell.column - 1 for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value}
    cards = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        card_id = row[headers["Id"]]
        if card_id and str(card_id).startswith(("Rocket", "Irene", "Zhouzhou")) and row[0] != "##":
            cards[str(card_id)] = row

    stale = [
        card_id for card_id, row in cards.items()
        if row[headers["CardType"]] == "Swift" and int(row[headers["ExecutingCost"]] or 0) != 0
    ]
    if stale:
        raise RuntimeError(f"Swift cards still have ExecutingCost: {stale}")

    for card_id in STARTER_OR_TEMPORARY_IDS:
        if str(cards[card_id][headers["IsInUpgrade"]]).lower() not in {"false", "0"}:
            raise RuntimeError(f"Starter/temporary card is still rewardable: {card_id}")

    for role in ("Rocket", "Irene", "Zhouzhou"):
        rarities = {row[headers["Rarity"]] for card_id, row in cards.items() if card_id.startswith(role)}
        if not {"普通", "稀有", "史诗"}.issubset(rarities):
            raise RuntimeError(f"{role} is missing rarity tiers: {rarities}")


def main() -> None:
    apply_card_updates()
    apply_character_traits()
    validate_written_workbooks()
    print("Character card audit fixes applied and validated.")


if __name__ == "__main__":
    main()
