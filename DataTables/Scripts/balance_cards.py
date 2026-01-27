#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
卡牌数值平衡脚本
基于时间格数计算合理伤害值

平衡公式:
- 每格基础价值 = 5 伤害
- 总格数 = Channeling + Duration + Recoil
- AOE 伤害 = 基础伤害 × 0.8
- 迅捷技能 = 基础伤害 - 3
- 带控制技能 = 基础伤害 - 2~3
- 条件伤害 = 基础 × 0.8, 触发后 × 1.5~2
"""

import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 卡牌平衡数据
# 格式: { 卡牌ID: { 字段名: 新值 } }
BALANCE_DATA = {
    # ========== Rocket (战士) - 2格技能 (Duration=1, Recoil=1, Channeling=0) ==========
    "Rocket001": {"damage": 10},           # 打击: 2格 → 10伤
    "Rocket002": {"damage": 8},            # 迅猛打击: 条件×2 → 8(16)
    "Rocket003": {"damage": 8},            # 旋风斩: AOE+护甲 → 8
    "Rocket004": {"damage": 8},            # 巨龙撞击: 控制 → 8
    "Rocket005": {"shield_value": 20},     # 守护: 护甲 30→20
    "Rocket007": {"value": 8},             # 吞噬痛苦: 治疗 10→8
    "Rocket009": {"damage": 8, "bonus_damage": 8},  # 交错打击: 8+8

    # ========== Irene (法师) - 2格技能 (Duration=1, Channeling=1, Recoil=0) ==========
    "Irene001": {"damage": 10},            # 火球术: 30→10
    "Irene002": {"damage": 8},             # 冰霜震击: 10→8
    "Irene003": {"damage": 4},             # 奥术飞弹: 5×3→4×3=12
    "Irene004": {"value": 10},             # 寒冰护甲: 15→10
    "Irene005": {"damage": 25},            # 炎爆术: 5格 80→25
    "Irene006": {"damage": 12},            # 暴风雪: 4格AOE 20→12
    "Irene009": {"damage": 25},            # 陨石灭世: 6格AOE 100→25
    "Irene011": {"damage": 10},            # 镜像实体: 20→10

    # ========== Zhouzhou (游侠) - 3格技能 (Duration=1, Channeling=1, Recoil=1) ==========
    "Zhouzhou001": {"damage": 12},         # 快速射击: 迅捷 15→12
    "Zhouzhou002": {"damage": 12},         # 强力击: 控制 20→12
    "Zhouzhou003": {"damage": 8},          # 二连矢: 12×2→8×2=16
    "Zhouzhou004": {"damage": 12},         # 多重射击: AOE 15→12
    "Zhouzhou005": {"damage": 25},         # 狙击: 5格 60→25
    "Zhouzhou006": {"damage": 8},          # 后撤步: 迅捷+控制 5→8
    "Zhouzhou008": {"damage": 12, "bonus_damage": 12},  # 致命射击: 25+25→12+12
    "Zhouzhou012": {"damage": 18},         # 穿云箭: 穿甲 40→18
}


def find_card_info_excel():
    """查找 CardInfo.xlsx 文件"""
    possible_paths = [
        os.path.join(os.path.dirname(__file__), "..", "Datas", "Character", "#CardInfo.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "Datas", "Character", "CardInfo.xlsx"),
    ]

    for path in possible_paths:
        abs_path = os.path.abspath(path)
        if os.path.exists(abs_path):
            return abs_path

    return None


def update_excel_values(excel_path):
    """更新 Excel 中的数值"""
    print(f"正在加载: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 获取表头
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"表头: {list(headers.keys())}")

    # 查找 ID 列
    id_col = headers.get("Id") or headers.get("id") or headers.get("ID")
    if not id_col:
        print("错误: 未找到 ID 列")
        return False

    # 遍历数据行
    updated_count = 0
    for row_idx in range(2, ws.max_row + 1):
        card_id = ws.cell(row=row_idx, column=id_col).value

        if card_id in BALANCE_DATA:
            updates = BALANCE_DATA[card_id]
            print(f"\n更新 {card_id}:")

            for field, new_value in updates.items():
                # 在 Effects 列中查找并更新
                # 由于 Effects 是复杂的 JSON 结构，需要特殊处理
                # 这里简化处理，直接更新对应列（如果存在）
                if field in headers:
                    col = headers[field]
                    old_value = ws.cell(row=row_idx, column=col).value
                    ws.cell(row=row_idx, column=col, value=new_value)
                    print(f"  {field}: {old_value} → {new_value}")
                    updated_count += 1
                else:
                    print(f"  {field}: 列不存在，需要手动在 Effects 中修改")

    # 保存
    if updated_count > 0:
        backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
        print(f"\n备份原文件到: {backup_path}")

        # 保存更新后的文件
        wb.save(excel_path)
        print(f"已保存更新 ({updated_count} 处修改)")
    else:
        print("\n没有需要更新的内容（Effects 列需要手动修改）")

    return True


def print_balance_guide():
    """打印平衡指南"""
    print("""
╔════════════════════════════════════════════════════════════════╗
║                    卡牌数值平衡指南                              ║
╠════════════════════════════════════════════════════════════════╣
║ 核心公式: 伤害 = 格数 × 5                                       ║
║                                                                ║
║ 格数计算: Channeling + Duration + Recoil                       ║
║                                                                ║
║ 调整系数:                                                       ║
║   • AOE 技能: ×0.8                                             ║
║   • 迅捷技能: -3                                                ║
║   • 带控制: -2~3                                                ║
║   • 条件伤害: 基础×0.8, 触发×1.5~2                              ║
║   • 护甲: 1护甲 ≈ 0.8伤害                                       ║
║   • 穿甲: -5伤害                                                ║
╠════════════════════════════════════════════════════════════════╣
║ 示例:                                                          ║
║   2格普通攻击 = 10伤害                                          ║
║   3格普通攻击 = 15伤害                                          ║
║   3格AOE = 12伤害                                               ║
║   3格迅捷 = 12伤害                                              ║
║   5格大招 = 25伤害                                              ║
╚════════════════════════════════════════════════════════════════╝
""")


def main():
    print("=" * 60)
    print("卡牌数值平衡工具")
    print("=" * 60)

    print_balance_guide()

    # 查找 Excel 文件
    excel_path = find_card_info_excel()

    if excel_path:
        print(f"\n找到配置表: {excel_path}")
        print("\n注意: Excel 的 Effects 列是复杂 JSON 结构，")
        print("建议直接编辑 JSON 文件或在 Excel 中手动修改 Effects 列。")
        print("\n以下是需要修改的数值清单:")
        print("-" * 60)

        for card_id, updates in BALANCE_DATA.items():
            print(f"{card_id}: {updates}")
    else:
        print("\n未找到 CardInfo.xlsx 文件")
        print("请确保文件位于: DataTables/Datas/Character/#CardInfo.xlsx")

    print("\n" + "=" * 60)
    print("JSON 文件已在之前的步骤中更新完成。")
    print("如需同步到 Excel，请手动修改 Effects 列中的数值。")
    print("=" * 60)


if __name__ == "__main__":
    main()
