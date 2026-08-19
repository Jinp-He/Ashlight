#!/usr/bin/env python3
"""Write Zhouzhou's v3 movement/morale card pool into #CardInfo.xlsx.

The cards intentionally use existing Effect beans for atomic actions.  The few
multi-target and runtime-history combinations are resolved by ZhouzhouCardCommands
in the battle layer, keyed by their stable card ids.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CARD_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CardInfo.xlsx"
CHARACTER_PATH = ROOT / "DataTables" / "Datas" / "Character" / "#CharaterInfo.xlsx"

FIELDS = (
    "Name", "Description", "Effects", "ChargeStartEffects", "ChargeWhileEffects",
    "CardType", "BelongTo", "TargetType", "Rarity", "AssetPath", "IsLocked",
    "Energy", "ExecutingCost", "IsEthereal", "IsExhaust", "TargetZone", "CastZone", "IsInUpgrade",
)


def card(name, description, effects, target, rarity, energy, *, ethereal="FALSE", exhaust="FALSE", reward="TRUE"):
    return {
        "Name": name, "Description": description, "Effects": effects,
        "ChargeStartEffects": None, "ChargeWhileEffects": None,
        "CardType": "Swift", "BelongTo": "Zhouzhou", "TargetType": target, "Rarity": rarity,
        "AssetPath": None, "IsLocked": "FALSE", "Energy": energy, "ExecutingCost": 0,
        "IsEthereal": ethereal, "IsExhaust": exhaust,
        "TargetZone": "Any", "CastZone": "Any", "IsInUpgrade": reward,
    }


CARDS = {
    "Zhouzhou000": card("移动", "[移动]一名友方角色。", "MovePositionEffect,N,Toggle", "SingleAlly", "基础", 1, ethereal="TRUE", exhaust="TRUE", reward="FALSE"),
    "Zhouzhou001": card("飞羽", "造成{A}点伤害。", "AttackEffect,A,4,false", "SingleEnemy", "基础", 0, reward="FALSE"),
    "Zhouzhou002": card("翻滚", "[移动]；获得[闪避]{F}。", "MovePositionEffect,N,Toggle;BuffEffect,F,Dodge,1", "Self", "基础", 1, reward="FALSE"),
    "Zhouzhou003": card("疾步", "[移动]；将1张[飞刀]加入手牌。", "MovePositionEffect,N,Toggle;AddToHandEffect,N,Extra001,1", "Self", "基础", 1, reward="FALSE"),
    "Zhouzhou004": card("踏歌", "[移动]一名其他友方角色。", "MovePositionEffect,N,Toggle", "SingleAlly", "基础", 1, reward="FALSE"),
    "Zhouzhou005": card("挽袖同行", "使一名其他友方角色获得{D}点护甲。", "DefenseEffect,D,7,false", "SingleAlly", "基础", 1, reward="FALSE"),
    "Zhouzhou006": card("喝彩", "获得[士气]{F}。", "BuffEffect,F,Morale,1", "Self", "基础", 0, exhaust="TRUE", reward="FALSE"),
    "Zhouzhou007": card("游身", "获得{D}点护甲；若拥有[士气]，再获得[闪避]{F}。", "DefenseEffect,D,5,false;BuffConditionalEffect,F,Dodge,1,HasMorale", "Self", "基础", 1, reward="FALSE"),
    "Zhouzhou009": card("紧急避险", "[移动]周周所在排的所有友方角色。", "MoveRowEffect,N", "Self", "基础", 1, reward="FALSE"),
    "Zhouzhou010": card("稳住阵脚", "所有友方角色获得3点护甲；自身获得1层[士气]。", None, "AllAlly", "普通", 1),
    "Zhouzhou011": card("钩锁", "[移动]一名敌方角色；对其造成{A}点伤害。", "MovePositionEffect,N,Toggle;AttackEffect,A,6,false", "SingleEnemy", "普通", 1),
    "Zhouzhou012": card("回锋", "造成{A}点伤害；若这张牌消耗了[士气]，获得{D}点护甲。", "AttackEffect,A,5,false", "SingleEnemy", "普通", 0),
    "Zhouzhou013": card("行云", "这个回合中，你首次成功移动后，抽1张牌。", None, "Self", "普通", 1),
    "Zhouzhou014": card("折返", "获得{D}点护甲和[闪避]{F}。", "DefenseEffect,D,4,false;BuffEffect,F,Dodge,1", "Self", "稀有", 1),
    "Zhouzhou015": card("雁双飞", "依次[移动]自己和一名其他友方角色；两者各获得4点护甲。", None, "SingleAlly", "稀有", 1),
    "Zhouzhou016": card("隧穿效应", "这个回合中，每当一名友方角色成功移动，将1张[飞刀]加入手牌，最多触发3次。", None, "Self", "稀有", 1),
    "Zhouzhou017": card("铁蒺藜", "这个回合中，每当任意角色成功移动，随机对一名敌方角色造成4点伤害，最多触发3次。", None, "Self", "稀有", 1),
    "Zhouzhou018": card("飒沓流星", "这个回合中，接下来每当你成功移动，抽1张牌，最多触发2次；将1张[飞刀]加入手牌。", None, "Self", "稀有", 1),
    "Zhouzhou019": card("一鼓作气", "消耗自身所有[士气]；所有友方角色获得5点护甲，每消耗1层，额外获得2点护甲。", None, "AllAlly", "稀有", 2),
    "Zhouzhou020": card("旧步重寻", "这个回合中，你首次成功移动后，将1张[飞刀]加入手牌。", None, "Self", "稀有", 1),
    "Zhouzhou021": card("千里不留行", "这个回合中，你的移动牌费用变为0；前两次成功移动时，各将1张[飞刀]加入手牌。", None, "Self", "史诗", 2),
    "Zhouzhou022": card("十步杀一人", "造成8点伤害；这个回合中，你每成功移动过1次，重复1次，最多重复2次。", None, "SingleEnemy", "史诗", 0),
    "Zhouzhou023": card("满场飞花", "[过载]1。依次[移动]至多2名不同的友方角色；被移动的角色各获得1层[闪避]。", "OverloadEffect,F,1", "AllAlly", "史诗", 2),
}

REMOVED_CARD_IDS = {"Zhouzhou008"}

EXTRAS = {
    "Extra001": {
        "Name": "飞刀", "Description": "[闪回]。对一名敌方角色造成{A}点伤害。", "Effects": "AttackEffect,A,3,false",
        "TargetType": "SingleEnemy", "BelongTo": "Zhouzhou", "IsEthereal": "TRUE", "IsExhaust": "FALSE",
    },
    "Extra005": {
        "Name": "移步", "Description": "[移动]。", "Effects": "MovePositionEffect,N,Toggle",
        "TargetType": "Self", "BelongTo": "Zhouzhou", "IsEthereal": "TRUE", "IsExhaust": "TRUE",
    },
}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src, dst = ws.cell(source_row, col), ws.cell(target_row, col)
        if src.has_style: dst._style = copy(src._style)
        if src.alignment: dst.alignment = copy(src.alignment)
        if src.protection: dst.protection = copy(src.protection)
        dst.number_format = src.number_format
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def find_rows(ws, id_col: int) -> dict[str, int]:
    return {str(ws.cell(row, id_col).value): row for row in range(5, ws.max_row + 1) if ws.cell(row, id_col).value}


def remove_cards(ws, card_ids: set[str]) -> None:
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    id_col = headers["Id"]
    for row in range(ws.max_row, 4, -1):
        if ws.cell(row, id_col).value in card_ids:
            ws.delete_rows(row, 1)


def write_cards(ws, cards: dict[str, dict], *, extra: bool) -> None:
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    rows = find_rows(ws, headers["Id"])
    for card_id, values in cards.items():
        row = rows.get(card_id)
        if row is None:
            row = ws.max_row + 1
            copy_row_style(ws, 5, row)
        ws.cell(row, 1).value = None
        ws.cell(row, headers["Id"]).value = card_id
        if extra:
            payload = {
                "CardType": "Swift", "Rarity": "临时", "AssetPath": None, "IsLocked": "FALSE", "Energy": 0,
                "ExecutingCost": 0, "TargetZone": "Any", "CastZone": "Any", "IsInUpgrade": "FALSE",
                "ChargeStartEffects": None, "ChargeWhileEffects": None,
            } | values
        else:
            payload = values
        for field in FIELDS:
            ws.cell(row, headers[field]).value = payload.get(field)
        ws.cell(row, headers["Effects"]).comment = None


def main() -> None:
    workbook = load_workbook(CARD_PATH)
    remove_cards(workbook["Rocket"], REMOVED_CARD_IDS)
    write_cards(workbook["Rocket"], CARDS, extra=False)
    write_cards(workbook["Extra"], EXTRAS, extra=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(CARD_PATH)

    character_book = load_workbook(CHARACTER_PATH)
    sheet = character_book["Sheet1"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row, headers["Character"]).value == "Zhouzhou":
            sheet.cell(row, headers["BaseDeck"]).value = (
                "Zhouzhou001,Zhouzhou001,Zhouzhou001,Zhouzhou001,Zhouzhou001,"
                "Zhouzhou007,Zhouzhou007,Zhouzhou007,Zhouzhou007,Zhouzhou007"
            )
            break
    else:
        raise RuntimeError("Zhouzhou character row is missing")
    if character_book.calculation is not None:
        character_book.calculation.fullCalcOnLoad = True
        character_book.calculation.forceFullCalc = True
    character_book.save(CHARACTER_PATH)
    print(f"Configured Zhouzhou v3: {len(CARDS)} main cards and {len(EXTRAS)} extra cards")


if __name__ == "__main__":
    main()
