"""Exports the map tile-content weight Excel source into the runtime JSON resource."""

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SOURCE = ROOT / "DataTables" / "Datas" / "Map" / "#MapTileContentWeight.xlsx"
OUTPUT = ROOT / "Assets" / "Resources" / "Config" / "map_tile_content_weights.json"
FIRST_DATA_ROW = 5


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(1, column).value for column in range(2, sheet.max_column + 1)]
    rows = []
    for row_index in range(FIRST_DATA_ROW, sheet.max_row + 1):
        values = [sheet.cell(row_index, column).value for column in range(2, sheet.max_column + 1)]
        if not any(value is not None for value in values):
            continue
        if isinstance(values[0], str) and values[0].startswith("##"):
            continue
        rows.append({header: value for header, value in zip(headers, values) if header})

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({"Rows": rows}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
