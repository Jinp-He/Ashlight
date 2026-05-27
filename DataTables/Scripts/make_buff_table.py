"""Create #BuffInfo.xlsx and expand __enums__.xlsx with new BuffEnum entries + BuffPolarityEnum.

Run from any cwd. Idempotent: safe to re-run.
"""
import os
import sys
import io
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent  # F:/Ashlight/DataTables
DATAS = ROOT / "Datas"
ENUMS_XLSX = DATAS / "__enums__.xlsx"
BUFF_XLSX = DATAS / "#BuffInfo.xlsx"


# ---------- Buff data ----------
# (Id, Name, Description, IconPath, Polarity, DefaultDuration, MaxStack, RefreshOnReapply)
BUFFS = [
    ("Channeling",    "引导",    "正在引导 {T} 格，中途被打断则施法失败",          "",                                "Neutral", -1, 1, True),
    ("Recoil",        "僵直",    "僵直中，剩余 {T} 格无法行动",                  "",                                "Debuff",  -1, 1, True),
    ("ReduceDmg",     "减伤",    "受到的伤害降低 {V}%，剩余 {T} 回合",            "UI/Buff/Icon_ReduceDmg",          "Buff",     2, 1, True),
    ("Chill",         "冰冻",    "行动延迟，剩余 {T} 回合",                       "UI/Buff/Icon_Chill",              "Debuff",   2, 3, True),
    ("Stun",          "晕眩",    "无法行动，剩余 {T} 回合",                       "UI/Buff/Icon_Stun",               "Debuff",   1, 1, True),
    ("Vulnerable",    "易伤",    "受到的伤害提高 {V}%，剩余 {T} 回合",            "UI/Buff/Icon_Vulnerable",         "Debuff",   2, 1, True),
    ("ReduceChannel", "引导缩短","下一次引导时间缩短 {V} 格",                     "UI/Buff/Icon_ReduceChannel",      "Buff",     1, 1, True),
    ("Root",          "禁锢",    "无法移动，剩余 {T} 回合",                       "UI/Buff/Icon_Root",               "Debuff",   2, 1, True),
    ("DrawCard",      "灵感",    "下回合开始时多抽 {V} 张牌",                     "UI/Buff/Icon_DrawCard",           "Buff",     1, 3, False),
    ("Taunt",         "嘲讽",    "敌人优先攻击你",                                "UI/Buff/Icon_Taunt",              "Buff",     2, 1, True),
    ("Taunted",       "被嘲讽",  "必须攻击发动嘲讽的角色",                        "UI/Buff/Icon_Taunted",            "Debuff",   2, 1, True),
    ("Stagger",       "破韧",    "累计承受 {V} 点伤害将打断行动",                 "UI/Buff/Icon_Stagger",            "Debuff",  -1, 1, True),
    ("Block",         "格挡",    "受到 {V} 次伤害将打断行动",                     "UI/Buff/Icon_Block",              "Debuff",  -1, 1, True),
    ("Strength",      "力量",    "攻击伤害 +{V}，剩余 {T} 回合",                  "UI/Buff/Icon_Strength",           "Buff",     2, 1, True),
    ("Weak",          "虚弱",    "造成的伤害降低 {V}%，剩余 {T} 回合",            "UI/Buff/Icon_Weak",               "Debuff",   2, 1, True),
    ("Frail",         "脆弱",    "获得的护甲降低 {V}%，剩余 {T} 回合",            "UI/Buff/Icon_Frail",              "Debuff",   2, 1, True),
    ("Dexterity",     "敏捷",    "防御牌额外 +{V} 护甲，剩余 {T} 回合",           "UI/Buff/Icon_Dexterity",          "Buff",     2, 1, True),
    ("Regen",         "再生",    "每回合开始回复 {V} 点生命，剩余 {T} 回合",      "UI/Buff/Icon_Regen",              "Buff",     3, 1, True),
    ("Energized",     "充能",    "下回合开始 +{V} 能量",                          "UI/Buff/Icon_Energized",          "Buff",     1, 3, False),
    ("CardCost",      "减费",    "下一张牌费用 -{V}",                             "UI/Buff/Icon_CardCost",           "Buff",     1, 1, True),
    ("Poison",        "中毒",    "每回合开始受到 {V} 点伤害，{V} 随之衰减 1",     "UI/Buff/Icon_Poison",             "Debuff",  -1, 99, False),
    ("Burn",          "燃烧",    "每回合开始受到 {V} 点伤害，剩余 {T} 回合",      "UI/Buff/Icon_Burn",               "Debuff",   3, 1, True),
    ("Artifact",      "圣物",    "抵消下一次 debuff，剩余 {V} 次",                "UI/Buff/Icon_Artifact",           "Buff",    -1, 99, False),
]


# ---------- Create #BuffInfo.xlsx ----------
def make_buff_info():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Schema rows (matches existing #CardInfo.xlsx convention)
    headers = ["##var", "Id", "Name", "Description", "IconPath", "Polarity",
               "DefaultDuration", "MaxStack", "RefreshOnReapply"]
    types = ["##type", "string", "string", "string", "string?", "BuffPolarityEnum",
             "int", "int", "bool"]
    groups = ["##group", "", "", "", "", "", "", "", ""]
    sep = ["##", "", "", "", "", "", "", "", ""]

    ws.append(headers)
    ws.append(types)
    ws.append(groups)
    ws.append(sep)

    for buff in BUFFS:
        ws.append(["", *buff])

    # Light styling: bold header, column widths
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=2, column=col_idx).font = Font(italic=True, color="888888")

    widths = [10, 16, 12, 50, 32, 16, 16, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "B5"

    wb.save(BUFF_XLSX)
    print(f"[OK] Wrote {BUFF_XLSX}")


# ---------- Update __enums__.xlsx ----------
def update_enums():
    wb = openpyxl.load_workbook(ENUMS_XLSX)
    ws = wb["Sheet1"]

    # 1) Find BuffEnum block (col B = "BuffEnum") and remove old items
    buff_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "BuffEnum":
            buff_row = r
            break
    if buff_row is None:
        raise RuntimeError("BuffEnum row not found in __enums__.xlsx")

    # Detect existing item rows: starting at buff_row, count consecutive rows where col H has a value
    end_row = buff_row
    while end_row <= ws.max_row and ws.cell(row=end_row, column=8).value:
        end_row += 1
    existing_count = end_row - buff_row  # number of item rows already present

    target_count = len(BUFFS)
    delta = target_count - existing_count

    if delta > 0:
        # Insert (delta) empty rows AFTER the last existing item, before next block
        ws.insert_rows(end_row, amount=delta)
    elif delta < 0:
        # Delete excess rows
        ws.delete_rows(end_row + delta, amount=-delta)

    # 2) Write all 11 items into col H (name) and col I (alias)
    #    First item row keeps full_name in col B; rest leave col B empty
    for i, buff in enumerate(BUFFS):
        r = buff_row + i
        bid, name, *_ = buff
        if i == 0:
            ws.cell(row=r, column=2, value="BuffEnum")
            ws.cell(row=r, column=3, value=False)   # flags
            ws.cell(row=r, column=4, value=True)    # unique
        else:
            ws.cell(row=r, column=2, value=None)
        ws.cell(row=r, column=8, value=bid)         # name
        ws.cell(row=r, column=9, value=name)        # alias (中文显示名)

    # 3) Append BuffPolarityEnum at the end (if not already present)
    have_polarity = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "BuffPolarityEnum":
            have_polarity = True
            break

    if not have_polarity:
        # Find first empty trailing row, then add a blank separator + block
        last = ws.max_row
        # Trim trailing empty rows to find true content end
        while last > 1 and all(ws.cell(row=last, column=c).value in (None, "") for c in range(1, 13)):
            last -= 1
        start = last + 2  # one blank line separator
        ws.cell(row=start, column=2, value="BuffPolarityEnum")
        ws.cell(row=start, column=3, value=False)
        ws.cell(row=start, column=4, value=True)
        ws.cell(row=start, column=8, value="Buff")
        ws.cell(row=start, column=9, value="增益")
        ws.cell(row=start + 1, column=8, value="Debuff")
        ws.cell(row=start + 1, column=9, value="减益")
        ws.cell(row=start + 2, column=8, value="Neutral")
        ws.cell(row=start + 2, column=9, value="中性")

    wb.save(ENUMS_XLSX)
    print(f"[OK] Updated {ENUMS_XLSX}")


if __name__ == "__main__":
    make_buff_info()
    update_enums()
    print("Done.")
