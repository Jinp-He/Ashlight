"""Configure the shared and character-theme UpgradeOptions pools.

The workbook is the Luban source of truth. Run this script from the project
root, then run Luban generation.
"""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = PROJECT_ROOT / "DataTables" / "Datas" / "Character" / "#UpgradeOptions.xlsx"

HEADERS = [
    "Id",
    "Name",
    "BelongTo",
    "IsCommon",
    "Theme",
    "Description",
    "Rarity",
    "Prerequisite",
    "Effects",
]

TYPES = [
    "string",
    "string",
    "CharacterEnum?",
    "bool",
    "string",
    "string",
    "RarityEnum",
    "string?",
    "(list#sep=;),(UpgradeEffect#sep=,)",
]

COMMENTS = [
    "升级选项Id(主键)",
    "升级选项名称",
    "从属角色(公共选项留空)",
    "是否进入所有角色的公共池",
    "构筑主题；副选项必须与前置主题一致",
    "升级文本",
    "稀有度",
    "前置升级选项(填Id,空=无)",
    "实装效果(多态;分隔,空=暂无)",
]

ROWS = [
    # 公共池：BelongTo 留空，由 IsCommon 决定对所有角色开放。
    ("UP_Common_001", "强健根基", None, True, "Common", "最大生命 +6", "Normal", None,
     "ModifyUnitStat,MaxHp,6,Self"),
    ("UP_Common_002", "战斗专注", None, True, "Common", "战斗开始时获得 1 点力量", "Normal", None,
     "GrantBuff,Strength,1,Self"),
    ("UP_Common_003", "灵巧身法", None, True, "Common", "战斗开始时获得 1 点敏捷", "Normal", None,
     "GrantBuff,Dexterity,1,Self"),

    # Rocket：保留旧挥砍链，并新增装甲、过载两条主题链。
    ("UP_Rocket_001", "强化挥砍", "Rocket", False, "DirectAttack", "「挥砍」造成的伤害 +5", "Normal", None,
     "ModifyCardStat,Rocket001,Damage,5"),
    ("UP_Rocket_002", "横扫挥砍", "Rocket", False, "DirectAttack", "「挥砍」改为对全体敌人造成伤害", "Rare",
     "UP_Rocket_001", "ModifyCardFlag,Rocket001,IsAoe,true"),
    ("UP_Rocket_003", "复合装甲", "Rocket", False, "Armor", "所有护甲牌的护甲 +2", "Normal", None,
     "ModifyCardStat,All,Defense,2"),
    ("UP_Rocket_004", "战线加固", "Rocket", False, "Armor",
     "「举盾」「装甲齐射」「阵线加固」「应急装甲」「区域装甲」「推进阵线」的护甲额外 +2", "Rare",
     "UP_Rocket_003",
     "ModifyCardStat,Rocket005,Defense,2;ModifyCardStat,Rocket011,Defense,2;"
     "ModifyCardStat,Rocket012,Defense,2;ModifyCardStat,Rocket016,Defense,2;"
     "ModifyCardStat,Rocket019,Defense,2;ModifyCardStat,Rocket020,Defense,2"),
    ("UP_Rocket_005", "过载弹头", "Rocket", False, "Overload",
     "「超载火炮」「震荡冲击」造成的伤害 +3", "Normal", None,
     "ModifyCardStat,Rocket004,Damage,3;ModifyCardStat,Rocket017,Damage,3"),
    ("UP_Rocket_006", "临界超载", "Rocket", False, "Overload",
     "「超载火炮」「震荡冲击」造成的伤害额外 +4", "Rare", "UP_Rocket_005",
     "ModifyCardStat,Rocket004,Damage,4;ModifyCardStat,Rocket017,Damage,4"),

    # Irene：执行牌与天气各形成一条“主选项 -> 同主题副选项”链。
    ("UP_Irene_001", "凝神", "Irene", False, "General", "战斗开始时获得 2 点力量", "Normal", None,
     "GrantBuff,Strength,2,Self"),
    ("UP_Irene_002", "咏唱增幅", "Irene", False, "Execution", "所有执行牌造成的伤害 +2", "Normal", None,
     "ModifyCardStat,CardType:Execution,Damage,2"),
    ("UP_Irene_003", "长线蓄能", "Irene", False, "Execution", "所有执行牌造成的伤害额外 +3", "Rare",
     "UP_Irene_002", "ModifyCardStat,CardType:Execution,Damage,3"),
    ("UP_Irene_004", "雷云导体", "Irene", False, "Weather", "「连锁闪电」造成的伤害 +4", "Normal", None,
     "ModifyCardStat,Irene009,Damage,4"),
    ("UP_Irene_005", "风暴共鸣", "Irene", False, "Weather",
     "「连锁闪电」造成的伤害额外 +6；「终焉倒数」造成的伤害 +4", "Rare", "UP_Irene_004",
     "ModifyCardStat,Irene009,Damage,6;ModifyCardStat,Irene020,Damage,4"),

    # Zhouzhou：移动带来的防护、飞刀收益各形成一条主题链。
    ("UP_Zhouzhou_001", "坚守", "Zhouzhou", False, "General", "最大生命 +8", "Epic", None,
     "ModifyUnitStat,MaxHp,8,Self"),
    ("UP_Zhouzhou_002", "流动掩护", "Zhouzhou", False, "Movement",
     "「挽袖同行」「游身」提供的护甲 +2", "Normal", None,
     "ModifyCardStat,Zhouzhou005,Defense,2;ModifyCardStat,Zhouzhou007,Defense,2"),
    ("UP_Zhouzhou_003", "无痕折返", "Zhouzhou", False, "Movement",
     "「挽袖同行」「游身」提供的护甲额外 +2", "Rare", "UP_Zhouzhou_002",
     "ModifyCardStat,Zhouzhou005,Defense,2;ModifyCardStat,Zhouzhou007,Defense,2"),
    ("UP_Zhouzhou_004", "飞刃随行", "Zhouzhou", False, "Movement", "「飞刀」造成的伤害 +2", "Normal", None,
     "ModifyCardStat,Extra001,Damage,2"),
    ("UP_Zhouzhou_005", "连步追刃", "Zhouzhou", False, "Movement",
     "「飞刀」造成的伤害额外 +2；「钩锁」造成的伤害 +2", "Rare", "UP_Zhouzhou_004",
     "ModifyCardStat,Extra001,Damage,2;ModifyCardStat,Zhouzhou011,Damage,2"),
]


def validate_rows() -> None:
    by_id = {}
    for record in ROWS:
        option_id = record[0]
        if option_id in by_id:
            raise RuntimeError(f"Duplicate upgrade option Id: {option_id}")
        by_id[option_id] = record

    for record in ROWS:
        option_id, _, belong_to, is_common, theme, _, _, prerequisite, _ = record
        if is_common and belong_to is not None:
            raise RuntimeError(f"Common option must not set BelongTo: {option_id}")
        if not is_common and belong_to is None:
            raise RuntimeError(f"Character option must set BelongTo: {option_id}")
        if not prerequisite:
            continue

        parent = by_id.get(prerequisite)
        if parent is None:
            raise RuntimeError(f"Missing prerequisite {prerequisite} for {option_id}")
        if parent[3] != is_common or parent[2] != belong_to or parent[4] != theme:
            raise RuntimeError(
                f"Prerequisite pool/owner/theme mismatch: {option_id} -> {prerequisite}"
            )


def configure() -> None:
    validate_rows()
    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook.active

    current_headers = [sheet.cell(1, column).value for column in range(2, sheet.max_column + 1)]
    if current_headers == ["Id", "Name", "BelongTo", "Description", "Rarity", "Prerequisite", "Effects"]:
        sheet.insert_cols(5, amount=2)
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row, 5)._style = copy(sheet.cell(row, 4)._style)
            sheet.cell(row, 6)._style = copy(sheet.cell(row, 7)._style)
    elif current_headers != HEADERS:
        raise RuntimeError(f"Unexpected UpgradeOptions headers: {current_headers}")

    # Clear old data but retain the established row formatting.
    for row in range(5, max(sheet.max_row, 5) + 1):
        for column in range(2, 11):
            sheet.cell(row, column).value = None

    for index, value in enumerate(HEADERS, start=2):
        sheet.cell(1, index).value = value
    for index, value in enumerate(TYPES, start=2):
        sheet.cell(2, index).value = value
    for index, value in enumerate(COMMENTS, start=2):
        sheet.cell(4, index).value = value

    template_row = 5
    for row_index, record in enumerate(ROWS, start=5):
        if row_index != template_row:
            for column in range(1, 11):
                sheet.cell(row_index, column)._style = copy(sheet.cell(template_row, column)._style)
                sheet.cell(row_index, column).number_format = sheet.cell(template_row, column).number_format
        for column, value in enumerate(record, start=2):
            sheet.cell(row_index, column).value = value

    # Keep the source deterministic and ask Excel-compatible tools to recalculate if needed.
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(WORKBOOK_PATH)


if __name__ == "__main__":
    configure()
