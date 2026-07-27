#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add Zhouzhou's Morale buff and movement trait to the Luban source workbooks.

The update is idempotent and preserves unrelated workbook rows and formatting.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DATAS = ROOT / "DataTables" / "Datas"
BUFF_PATH = DATAS / "#BuffInfo.xlsx"
ENUMS_PATH = DATAS / "__enums__.xlsx"
CHARACTER_PATH = DATAS / "Character" / "#CharaterInfo.xlsx"
BACKUP_DIR = ROOT / "DataTables" / "_backups"

MORALE_ROW = {
    "Id": "Morale",
    "Name": "士气",
    "Description": "特殊增益，当前 {V} 层",
    "IconPath": "UI/Buff/Icon_Morale",
    "Polarity": "Buff",
    "DefaultDuration": -1,
    "MaxStack": 3,
    "RefreshOnReapply": False,
}
ZHOUZHOU_TRAIT = "MoveAllyGrantMorale"


def backup_once(path: Path) -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"{path.stem}_backup_before_morale.xlsx"
    if not backup.exists():
        shutil.copy2(path, backup)


def headers_for(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def update_buff_info() -> None:
    backup_once(BUFF_PATH)
    wb = load_workbook(BUFF_PATH)
    ws = wb["Sheet1"]
    headers = headers_for(ws)

    target_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=headers["Id"]).value == MORALE_ROW["Id"]:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1

    for field, value in MORALE_ROW.items():
        ws.cell(row=target_row, column=headers[field], value=value)

    wb.save(BUFF_PATH)


def update_buff_enum() -> None:
    backup_once(ENUMS_PATH)
    wb = load_workbook(ENUMS_PATH)
    ws = wb["Sheet1"]

    buff_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == "BuffEnum":
            buff_row = row
            break
    if buff_row is None:
        raise RuntimeError("BuffEnum block not found")

    end_row = buff_row
    existing = {}
    while end_row <= ws.max_row and ws.cell(row=end_row, column=8).value:
        existing[str(ws.cell(row=end_row, column=8).value)] = end_row
        end_row += 1

    additions = [("Resolve", "坚毅"), ("Morale", "士气")]
    missing = [(name, alias) for name, alias in additions if name not in existing]
    if missing:
        ws.insert_rows(end_row, amount=len(missing))
        for offset, (name, alias) in enumerate(missing):
            row = end_row + offset
            ws.cell(row=row, column=8, value=name)
            ws.cell(row=row, column=9, value=alias)

    wb.save(ENUMS_PATH)


def update_character_trait() -> None:
    backup_once(CHARACTER_PATH)
    wb = load_workbook(CHARACTER_PATH)
    ws = wb["Sheet1"]
    headers = headers_for(ws)

    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=headers["Character"]).value == "Zhouzhou":
            ws.cell(row=row, column=headers["Trait"], value=ZHOUZHOU_TRAIT)
            wb.save(CHARACTER_PATH)
            return
    raise RuntimeError("Zhouzhou row not found in #CharaterInfo.xlsx")


def validate() -> None:
    wb = load_workbook(BUFF_PATH, data_only=False, read_only=True)
    ws = wb["Sheet1"]
    headers = {cell.value: cell.column - 1 for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value}
    morale = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[headers["Id"]] == "Morale":
            morale = row
            break
    if morale is None or morale[headers["MaxStack"]] != 3 or morale[headers["DefaultDuration"]] != -1:
        raise RuntimeError("Morale BuffInfo validation failed")

    wb = load_workbook(CHARACTER_PATH, data_only=False, read_only=True)
    ws = wb["Sheet1"]
    headers = {cell.value: cell.column - 1 for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value}
    traits = {
        row[headers["Character"]]: row[headers["Trait"]]
        for row in ws.iter_rows(min_row=4, values_only=True)
        if row[headers["Character"]]
    }
    if traits.get("Zhouzhou") != ZHOUZHOU_TRAIT:
        raise RuntimeError("Zhouzhou Trait validation failed")


def main() -> None:
    update_buff_info()
    update_buff_enum()
    update_character_trait()
    validate()
    print("Zhouzhou Morale buff and movement trait configured.")


if __name__ == "__main__":
    main()
